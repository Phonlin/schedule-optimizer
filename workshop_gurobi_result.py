# In[1]:


import gurobipy as gp
from gurobipy import GRB
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

#INPUT DATA
eng_df    = pd.read_csv(r"Engineer_List.csv")
demand_df = pd.read_csv(r"Shift_Demand.csv")

NUM_DAYS = 30
NUM_ENG  = len(eng_df)
engineers  = eng_df['人員'].tolist()
default_group = dict(zip(eng_df['人員'], eng_df['班別群組(*第一碼為群組代碼第二碼之後為可backup群組)']))

#Pre-assigned shifts
pre_assigned = {}
date_cols = [f'Date_{d+1}' for d in range(NUM_DAYS)]
for i, row in eng_df.iterrows():
    for d, col in enumerate(date_cols):
        val = str(row[col]).strip() if pd.notna(row[col]) else ''
        if val in ('D', 'E', 'N', 'O'):
            pre_assigned.setdefault(i, {})[d] = val

demand_D = demand_df['Day'].tolist()
demand_E = demand_df['Afternoon'].tolist()
demand_N = demand_df['Night'].tolist()
is_weekend = [(str(v).strip() == 'Y') for v in demand_df['IfWeekend'].tolist()]
weekend_days = [d for d in range(NUM_DAYS) if is_weekend[d]]

SHIFTS    = ['D', 'E', 'N', 'O']
S_IDX     = {s: k for k, s in enumerate(SHIFTS)} 
I         = range(NUM_ENG)
D         = range(NUM_DAYS)
S         = range(len(SHIFTS))

#GUROBI MODEL
m = gp.Model("TSMC_Scheduling")
m.setParam('TimeLimit', 60)   # 1 min
m.setParam('MIPGap',    0.01)  # 1 % optimality gap
m.setParam('OutputFlag', 1)

#Decision variables
x = m.addVars(NUM_ENG, NUM_DAYS, len(SHIFTS), vtype=GRB.BINARY, name="x")

#constraint 1: each engineer gets exactly 1 shift per day
for i in I:
    for d in D:
        m.addConstr(gp.quicksum(x[i, d, s] for s in S) == 1,
                    name=f"one_shift_{i}_{d}")

#constraint 2: pre-assigned shifts must be honoured
for i, days in pre_assigned.items():
    for d, sh in days.items():
        m.addConstr(x[i, d, S_IDX[sh]] == 1,
                    name=f"preassign_{i}_{d}")

#constraint 3: daily demand
for d in D:
    m.addConstr(gp.quicksum(x[i, d, S_IDX['D']] for i in I) == demand_D[d], name=f"dem_D_{d}")
    m.addConstr(gp.quicksum(x[i, d, S_IDX['E']] for i in I) == demand_E[d], name=f"dem_E_{d}")
    m.addConstr(gp.quicksum(x[i, d, S_IDX['N']] for i in I) == demand_N[d], name=f"dem_N_{d}")

#PENALTY VARIABLES & OBJECTIVE
obj_terms = []

#連續上班6天 
#建立二元變數=1 表示第i個人在第d天形成連續6天上班
z_consec6 = m.addVars(NUM_ENG, NUM_DAYS, vtype=GRB.BINARY, name="z_consec6")
for i in I:
    for d in range(5, NUM_DAYS):
        # 從第6天開始（因為要往前看6天d-5 ~ d）
        # 計算這6天（d-5 ~ d）每天是否有上班（非 O）
        # 每一天如果上班=1，休假=0
        working_expr = [gp.quicksum(x[i, d2, s] for s in S if SHIFTS[s] != 'O')
                        for d2 in range(d - 5, d + 1)]
        
        # 下界限制：
        # z>=(6天上班總數 - 5)
        # 若6天都上班 sum=6 → z >=1 → z=1（觸發違規）
        # 若少於6天 z >=0或負數 z可為0
        m.addConstr(z_consec6[i, d] >= gp.quicksum(working_expr) - 5,
                    name=f"consec6_lb_{i}_{d}")

        # 上界限制（逐天限制）：
        # z<=每一天是否上班
        # 只要6天中有一天休假 z <= 0 → z=0（不會誤判）
        for d2 in range(d - 5, d + 1):
            m.addConstr(z_consec6[i, d] <= gp.quicksum(x[i, d2, s]
                        for s in S if SHIFTS[s] != 'O'),
                        name=f"consec6_ub_{i}_{d}_{d2}")

        # 加入目標式（懲罰1）
        obj_terms.append(1.0 * z_consec6[i, d])

#不合法班別銜接（Illegal transitions）
z_trans = m.addVars(NUM_ENG, NUM_DAYS - 1, 3, vtype=GRB.BINARY, name="z_trans")

for i in I:
    for d in range(NUM_DAYS - 1):
        
        m.addConstr(z_trans[i, d, 0] >= x[i, d,   S_IDX['N']] +
                                         x[i, d+1, S_IDX['D']] - 1,
                    name=f"trans0a_{i}_{d}")
        m.addConstr(z_trans[i, d, 0] >= x[i, d,   S_IDX['N']] +
                                         x[i, d+1, S_IDX['E']] - 1,
                    name=f"trans0b_{i}_{d}")
        
        # Pattern 1: afternoon today, then day tomorrow
        m.addConstr(z_trans[i, d, 1] >= x[i, d,   S_IDX['E']] +
                                         x[i, d+1, S_IDX['D']] - 1,
                    name=f"trans1_{i}_{d}")
        
        # Pattern 2: day or afternoon today, then night tomorrow
        m.addConstr(z_trans[i, d, 2] >= x[i, d,   S_IDX['D']] +
                                         x[i, d+1, S_IDX['N']] - 1,
                    name=f"trans2a_{i}_{d}")
        m.addConstr(z_trans[i, d, 2] >= x[i, d,   S_IDX['E']] +
                                         x[i, d+1, S_IDX['N']] - 1,
                    name=f"trans2b_{i}_{d}")
        
        for k in range(3):
            obj_terms.append(1.0 * z_trans[i, d, k])

#違反預設班別 Violate default shift group
#二元變數 =1 表示第 i 個人在第 d 天排到「非預設班別且非休假」違規
z_viol_def = m.addVars(NUM_ENG, NUM_DAYS, vtype=GRB.BINARY, name="z_viol_def")

for i in I:
    grp = default_group[engineers[i]] # 取得員工的預設班別
    non_grp_non_O = [s for s in S if SHIFTS[s] != 'O' and SHIFTS[s] != grp]
    
    for d in D:
        for s in non_grp_non_O:
            m.addConstr(z_viol_def[i, d] >= x[i, d, s],
                        name=f"viol_def_{i}_{d}_{s}")
        obj_terms.append(0.2 * z_viol_def[i, d])

#月總休假 < 9 天
z_off_short = m.addVars(NUM_ENG, lb=0, vtype=GRB.INTEGER, name="z_off_short")

for i in I:
    total_off = gp.quicksum(x[i, d, S_IDX['O']] for d in D) # 計算員工整個月的休假天數

    # z >= (9 - total off)
    # 若休假不足 則z會變成不足的天數
    # 若休假 >=9 則RHS為負,z可為0
    m.addConstr(z_off_short[i] >= 9 - total_off, name=f"off_short_{i}")
    obj_terms.append(0.1 * z_off_short[i])

#周末休假 < 4 天
#建立整數變數表示周末休假不足天數
z_wk_short = m.addVars(NUM_ENG, lb=0, vtype=GRB.INTEGER, name="z_wk_short")
for i in I:
    wk_off = gp.quicksum(x[i, d, S_IDX['O']] for d in weekend_days) # 計算員工在周末的休假天數
    
    # 若周末休假 < 4 則 z = (4 - wk_off)
    m.addConstr(z_wk_short[i] >= 4 - wk_off, name=f"wk_short_{i}")
    obj_terms.append(0.1 * z_wk_short[i])

#僅排休1日
z_iso = m.addVars(NUM_ENG, NUM_DAYS, vtype=GRB.BINARY, name="z_iso")

for i in I:
    for d in range(1, NUM_DAYS - 1): # 不檢查第一天與最後一天（因為沒有前後）

        # 條件：天是休假 AND 前一天不是休假 AND 後一天不是休假
        m.addConstr(z_iso[i, d] >= x[i, d,   S_IDX['O']]
                                  - x[i, d-1, S_IDX['O']]
                                  - x[i, d+1, S_IDX['O']],
                    name=f"iso_{i}_{d}")
        
        obj_terms.append(0.1 * z_iso[i, d])

#連續休假區塊數 < 2
#z=1 表示第 d 天是休假區塊的起點
z_bstart = m.addVars(NUM_ENG, NUM_DAYS, vtype=GRB.BINARY, name="z_bstart")
for i in I:
    # 第一天特別處理
    # 若第0天是休假就是區塊起點
    m.addConstr(z_bstart[i, 0] == x[i, 0, S_IDX['O']],
                name=f"bstart0_{i}")
    
    for d in range(1, NUM_DAYS):
        m.addConstr(z_bstart[i, d] <= x[i, d,   S_IDX['O']], name=f"bstart_ub1_{i}_{d}")
        m.addConstr(z_bstart[i, d] <= 1 - x[i, d-1, S_IDX['O']], name=f"bstart_ub2_{i}_{d}")
        m.addConstr(z_bstart[i, d] >= x[i, d, S_IDX['O']] - x[i, d-1, S_IDX['O']],
                    name=f"bstart_lb_{i}_{d}")

# 建立變數：休假區塊不足數
z_consec_off_short = m.addVars(NUM_ENG, lb=0, vtype=GRB.INTEGER, name="z_consec_off_short")

# 計算員工總共有幾個休假區塊
for i in I:
    total_blocks = gp.quicksum(z_bstart[i, d] for d in D)

    # 若區塊數 < 2 則 z = (2 - blocks)
    m.addConstr(z_consec_off_short[i] >= 2 - total_blocks,
                name=f"consec_off_short_{i}")

    # 每少一個休假區塊, 罰0.1
    obj_terms.append(0.1 * z_consec_off_short[i])

m.setObjective(gp.quicksum(obj_terms), GRB.MINIMIZE)

# SOLVE

print("=" * 60)
print("TSMC Scheduling — Gurobi MIP Solver")
print("=" * 60)
m.optimize()

if m.Status not in (GRB.OPTIMAL, GRB.TIME_LIMIT, GRB.SUBOPTIMAL):
    print(f"No feasible solution found (status={m.Status}).")
    raise SystemExit(1)

print(f"\nObjective value (penalty): {m.ObjVal:.4f}")


schedule = []
for i in I:
    row = []
    for d in D:
        for s in S:
            if x[i, d, s].X > 0.5:
                row.append(SHIFTS[s])
                break
    schedule.append(row)

#COMPUTE PENALTY BREAKDOWN
def compute_breakdown(schedule):
    bd = {k: 0 for k in ['consecutive_6days','night_to_day_aft','aft_to_day','day_aft_to_night',
                          'violate_default','consec_off_lt2','total_off_lt9','weekend_off_lt4','isolated_off']}
    for i in range(NUM_ENG):
        s = schedule[i]; grp = default_group[engineers[i]]
        consec = 0
        for d in range(NUM_DAYS):
            if s[d] != 'O': consec += 1; bd['consecutive_6days'] += (1 if consec >= 6 else 0)
            else: consec = 0
        for d in range(NUM_DAYS - 1):
            t, n = s[d], s[d+1]
            if t == 'N' and n in ('D','E'): bd['night_to_day_aft'] += 1
            if t == 'E' and n == 'D':       bd['aft_to_day']       += 1
            if t in ('D','E') and n == 'N': bd['day_aft_to_night'] += 1
        for d in range(NUM_DAYS):
            if s[d] != 'O' and s[d] != grp: bd['violate_default'] += 1
        off_blocks = []; in_b = False; bl = 0
        for d in range(NUM_DAYS):
            if s[d] == 'O': in_b = True; bl += 1
            else:
                if in_b: off_blocks.append(bl); in_b = False; bl = 0
        if in_b: off_blocks.append(bl)
        coc = len(off_blocks)
        if coc < 2: bd['consec_off_lt2'] += (2 - coc)
        to = s.count('O')
        if to < 9: bd['total_off_lt9'] += (9 - to)
        wo = sum(1 for d in range(NUM_DAYS) if is_weekend[d] and s[d] == 'O')
        if wo < 4: bd['weekend_off_lt4'] += (4 - wo)
        for d in range(1, NUM_DAYS - 1):
            if s[d] == 'O' and s[d-1] != 'O' and s[d+1] != 'O': bd['isolated_off'] += 1
    return bd

breakdown = compute_breakdown(schedule)
raw_penalty = (breakdown['consecutive_6days'] * 1.0 +
               breakdown['night_to_day_aft']  * 1.0 +
               breakdown['aft_to_day']        * 1.0 +
               breakdown['day_aft_to_night']  * 1.0 +
               breakdown['violate_default']   * 0.2 +
               breakdown['consec_off_lt2']    * 0.1 +
               breakdown['total_off_lt9']     * 0.1 +
               breakdown['weekend_off_lt4']   * 0.1 +
               breakdown['isolated_off']      * 0.1)

print("\n PENALTY BREAKDOWN ")
print(f"  連續上班6天   (×1.0): cnt={breakdown['consecutive_6days']}  => {breakdown['consecutive_6days']*1.0:.1f}")
print(f"  夜→早/午      (×1.0): cnt={breakdown['night_to_day_aft']}  => {breakdown['night_to_day_aft']*1.0:.1f}")
print(f"  午→早         (×1.0): cnt={breakdown['aft_to_day']}  => {breakdown['aft_to_day']*1.0:.1f}")
print(f"  早/午→夜      (×1.0): cnt={breakdown['day_aft_to_night']}  => {breakdown['day_aft_to_night']*1.0:.1f}")
print(f"  違反預設班別  (×0.2): cnt={breakdown['violate_default']}  => {breakdown['violate_default']*0.2:.1f}")
print(f"  連續休假<2    (×0.1): cnt={breakdown['consec_off_lt2']}  => {breakdown['consec_off_lt2']*0.1:.1f}")
print(f"  月休<9天      (×0.1): cnt={breakdown['total_off_lt9']}  => {breakdown['total_off_lt9']*0.1:.1f}")
print(f"  周末休<4天    (×0.1): cnt={breakdown['weekend_off_lt4']}  => {breakdown['weekend_off_lt4']*0.1:.1f}")
print(f"  僅排休1天      (×0.1): cnt={breakdown['isolated_off']}  => {breakdown['isolated_off']*0.1:.1f}")
print(f"  ──────────────────────────────────────────")
print(f"  TOTAL PENALTY: {raw_penalty:.2f}")

#PER-ENGINEER STATS

per_eng_stats = []
for i in range(NUM_ENG):
    s = schedule[i]; grp = default_group[engineers[i]]
    consec_6 = 0; consec = 0
    for d in range(NUM_DAYS):
        if s[d] != 'O': consec += 1; consec_6 += (1 if consec >= 6 else 0)
        else: consec = 0
    night_da = sum(1 for d in range(NUM_DAYS-1) if s[d]=='N' and s[d+1] in ('D','E'))
    aft_d    = sum(1 for d in range(NUM_DAYS-1) if s[d]=='E' and s[d+1]=='D')
    da_night = sum(1 for d in range(NUM_DAYS-1) if s[d] in ('D','E') and s[d+1]=='N')
    viol_def = sum(1 for d in range(NUM_DAYS) if s[d]!='O' and s[d]!=grp)
    off_blocks=[]; in_b=False; bl=0
    for d in range(NUM_DAYS):
        if s[d]=='O': in_b=True; bl+=1
        else:
            if in_b: off_blocks.append(bl); in_b=False; bl=0
    if in_b: off_blocks.append(bl)
    coc=len(off_blocks); to=s.count('O')
    wo=sum(1 for d in range(NUM_DAYS) if is_weekend[d] and s[d]=='O')
    iso=sum(1 for d in range(1,NUM_DAYS-1) if s[d]=='O' and s[d-1]!='O' and s[d+1]!='O')
    p6=consec_6*1.0; pnd=night_da*1.0; pad=aft_d*1.0; pdn=da_night*1.0; pvd=viol_def*0.2
    pco=max(0,2-coc)*0.1; pto=max(0,9-to)*0.1; pwo=max(0,4-wo)*0.1; pio=iso*0.1
    pt=p6+pnd+pad+pdn+pvd+pco+pto+pwo+pio
    per_eng_stats.append(dict(engineer=engineers[i],group=grp,schedule=s,
        total_off=to,weekend_off=wo,consec_off_blocks=coc,isolated_off=iso,
        consec_6=consec_6,night_to_da=night_da,aft_to_d=aft_d,da_to_night=da_night,viol_default=viol_def,
        p_consec6=p6,p_night_da=pnd,p_aft_d=pad,p_da_night=pdn,p_viol_def=pvd,
        p_consec_off=pco,p_total_off=pto,p_wk_off=pwo,p_iso_off=pio,p_total=pt))

#EXCEL
wb  = Workbook()
ws1 = wb.active;                ws1.title = "Sheet1_排班結果"
ws2 = wb.create_sheet("Sheet2_懲罰統計")

COLOR_D    = "C6EFCE"; COLOR_E = "FFEB9C"; COLOR_N = "9DC3E6"; COLOR_O = "F2F2F2"
COLOR_HDR  = "1F4E79"; COLOR_HDR2 = "2E75B6"; COLOR_SUM = "E2EFDA"
thin       = Side(style='thin',   color='AAAAAA')
thk        = Side(style='medium', color='000000')
bdr        = Border(left=thin, right=thin, top=thin, bottom=thin)
bdr_thick  = Border(left=thk,  right=thk,  top=thk,  bottom=thk)

def cs(ws, row, col, val, bg=None, bold=False, fc="000000", fs=10, wrap=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name='Arial', bold=bold, color=fc, size=fs)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=wrap)
    c.border    = bdr
    if bg: c.fill = PatternFill("solid", start_color=bg, end_color=bg)
    return c

#Sheet 1
ws1.merge_cells(f'A1:{get_column_letter(NUM_DAYS+3)}1')
c = ws1.cell(1, 1, "TSMC Scheduling Output — 排班結果 (Gurobi MIP 最小化懲罰值)")
c.font = Font(name='Arial', bold=True, size=14, color="FFFFFF")
c.fill = PatternFill("solid", start_color=COLOR_HDR, end_color=COLOR_HDR)
c.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 28

cs(ws1, 2, 1, "人員",     COLOR_HDR2, bold=True, fc="FFFFFF")
cs(ws1, 2, 2, "班別群組", COLOR_HDR2, bold=True, fc="FFFFFF")
for d in range(NUM_DAYS):
    bg = "FFC000" if is_weekend[d] else COLOR_HDR2
    cs(ws1, 2, d+3, f"Date_{d+1}", bg, bold=True, fc="FFFFFF")
ws1.row_dimensions[2].height = 18

sc = {'D': COLOR_D, 'E': COLOR_E, 'N': COLOR_N, 'O': COLOR_O}
for i, stat in enumerate(per_eng_stats):
    row = i + 3
    nbg = "FFE0E0" if stat['p_total'] > 0 else "FFFFFF"
    cs(ws1, row, 1, stat['engineer'], nbg, bold=(stat['p_total'] > 0))
    cs(ws1, row, 2, stat['group'], "E2EFDA", bold=True)
    for d in range(NUM_DAYS):
        sh      = stat['schedule'][d]
        is_pre  = (i in pre_assigned and d in pre_assigned.get(i, {}))
        c       = ws1.cell(row=row, column=d+3, value=sh)
        c.font      = Font(name='Arial', size=9, bold=is_pre)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill      = PatternFill("solid", start_color=sc.get(sh, "FFFFFF"), end_color=sc.get(sh, "FFFFFF"))
        c.border    = bdr_thick if is_pre else bdr
    ws1.row_dimensions[row].height = 16

for label, row, color, arr in [
    ("需求_D", NUM_ENG+3, COLOR_D, demand_D),
    ("需求_E", NUM_ENG+4, COLOR_E, demand_E),
    ("需求_N", NUM_ENG+5, COLOR_N, demand_N)
]:
    ws1.merge_cells(f'A{row}:B{row}')
    c = ws1.cell(row=row, column=1, value=label)
    c.font = Font(name='Arial', bold=True, size=9)
    c.fill = PatternFill("solid", start_color=color, end_color=color)
    c.alignment = Alignment(horizontal='center', vertical='center')
    shift_key = label[-1]   # 'D', 'E', or 'N'
    for d in range(NUM_DAYS):
        cnt = sum(1 for ii in range(NUM_ENG) if schedule[ii][d] == shift_key)
        req = arr[d]; met = (cnt == req)
        c2 = ws1.cell(row=row, column=d+3, value=f"{cnt}/{req}")
        c2.font = Font(name='Arial', size=9, color="000000" if met else "FF0000", bold=(not met))
        c2.fill = PatternFill("solid", start_color=color, end_color=color)
        c2.alignment = Alignment(horizontal='center', vertical='center')
        c2.border = bdr

ws1.column_dimensions['A'].width = 14
ws1.column_dimensions['B'].width = 10
for d in range(NUM_DAYS):
    ws1.column_dimensions[get_column_letter(d+3)].width = 5.5

leg = NUM_ENG + 7
ws1.cell(leg, 1, "圖例：").font = Font(bold=True, name='Arial')
for sh, color, label, col in [('D',COLOR_D,'早班(D)',2),('E',COLOR_E,'午班(E)',3),
                                ('N',COLOR_N,'晚班(N)',4),('O',COLOR_O,'休假(O)',5)]:
    c = ws1.cell(leg, col, label)
    c.fill = PatternFill("solid", start_color=color, end_color=color)
    c.alignment = Alignment(horizontal='center')
    c.font = Font(name='Arial', size=9)
c = ws1.cell(leg, 6, "加粗框 = 預先指定班次")
c.font = Font(name='Arial', size=9, bold=True)
c = ws1.cell(leg, 8, "橘色欄標題 = 周末")
c.fill = PatternFill("solid", start_color="FFC000", end_color="FFC000")
c.font = Font(name='Arial', size=9)

# ── Sheet 2 ──────────────────────────────────
ws2.merge_cells('A1:P1')
c = ws2.cell(1, 1, "TSMC Scheduling — 懲罰值統計 (Penalty Score Summary) | Gurobi MIP")
c.font = Font(name='Arial', bold=True, size=14, color="FFFFFF")
c.fill = PatternFill("solid", start_color=COLOR_HDR, end_color=COLOR_HDR)
c.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 28

hdrs = ["人員","班別群組","總休假數","周末休假數","連續休假次數","僅排休1天次數",
        "連上6天(次)","夜→早/午(次)","午→早(次)","早/午→夜(次)","違反預設(次)",
        "懲罰:連上6天\n(×1.0)","懲罰:違法接班\n(×1.0)","懲罰:違反預設\n(×0.2)",
        "懲罰:公平性\n(×0.1)","合計懲罰值"]
for col, h in enumerate(hdrs, 1):
    cs(ws2, 2, col, h, COLOR_HDR2, bold=True, fc="FFFFFF", fs=9, wrap=True)
ws2.row_dimensions[2].height = 30

for i, stat in enumerate(per_eng_stats):
    row = i + 3
    bg  = "FFE0E0" if stat['p_total'] > 0 else "FFFFFF"
    ill  = stat['p_night_da'] + stat['p_aft_d'] + stat['p_da_night']
    fair = stat['p_consec_off'] + stat['p_total_off'] + stat['p_wk_off'] + stat['p_iso_off']
    vals = [stat['engineer'], stat['group'],
            stat['total_off'], stat['weekend_off'], stat['consec_off_blocks'], stat['isolated_off'],
            stat['consec_6'], stat['night_to_da'], stat['aft_to_d'], stat['da_to_night'], stat['viol_default'],
            round(stat['p_consec6'],2), round(ill,2), round(stat['p_viol_def'],2),
            round(fair,2), round(stat['p_total'],2)]
    for col, v in enumerate(vals, 1):
        is_p = col >= 12; is_t = col == 16
        cbg  = ("FFD700" if is_t else ("FFF2CC" if is_p else bg))
        if is_t: cbg = "FF8080" if stat['p_total'] > 0 else "C6EFCE"
        cs(ws2, row, col, v, cbg, bold=is_t, fs=9)
    ws2.row_dimensions[row].height = 16

sr = NUM_ENG + 3
cs(ws2, sr, 1, "合計", COLOR_SUM, bold=True); cs(ws2, sr, 2, "", COLOR_SUM)
sk  = ['total_off','weekend_off','consec_off_blocks','isolated_off','consec_6',
       'night_to_da','aft_to_d','da_to_night','viol_default']
ss  = {k: sum(st[k] for st in per_eng_stats) for k in sk}
sp  = {k: sum(st[k] for st in per_eng_stats)
       for k in ['p_consec6','p_night_da','p_aft_d','p_da_night','p_viol_def',
                 'p_consec_off','p_total_off','p_wk_off','p_iso_off','p_total']}
ill_t  = sp['p_night_da'] + sp['p_aft_d'] + sp['p_da_night']
fair_t = sp['p_consec_off'] + sp['p_total_off'] + sp['p_wk_off'] + sp['p_iso_off']
sv = ["","", ss['total_off'],ss['weekend_off'],ss['consec_off_blocks'],ss['isolated_off'],
      ss['consec_6'],ss['night_to_da'],ss['aft_to_d'],ss['da_to_night'],ss['viol_default'],
      round(sp['p_consec6'],2), round(ill_t,2), round(sp['p_viol_def'],2),
      round(fair_t,2), round(sp['p_total'],2)]
for col, v in enumerate(sv, 1):
    cs(ws2, sr, col, v, COLOR_SUM, bold=True, fs=9)
ws2.row_dimensions[sr].height = 18

# Penalty breakdown summary table
br = NUM_ENG + 6
bdata = [
    ("懲罰項目",        "權重", "違反次數",                    "懲罰值"),
    ("連續上班≥6天",    1.0,    breakdown['consecutive_6days'], round(breakdown['consecutive_6days']*1.0, 2)),
    ("夜班→早/午班",    1.0,    breakdown['night_to_day_aft'],  round(breakdown['night_to_day_aft']*1.0,  2)),
    ("午班→早班",       1.0,    breakdown['aft_to_day'],        round(breakdown['aft_to_day']*1.0,        2)),
    ("早/午班→夜班",    1.0,    breakdown['day_aft_to_night'],  round(breakdown['day_aft_to_night']*1.0,  2)),
    ("違反預設班別",    0.2,    breakdown['violate_default'],   round(breakdown['violate_default']*0.2,   2)),
    ("連續休假次數<2",  0.1,    breakdown['consec_off_lt2'],    round(breakdown['consec_off_lt2']*0.1,    2)),
    ("月休假<9天",      0.1,    breakdown['total_off_lt9'],     round(breakdown['total_off_lt9']*0.1,     2)),
    ("周末休假<4天",    0.1,    breakdown['weekend_off_lt4'],   round(breakdown['weekend_off_lt4']*0.1,   2)),
    ("僅排休1天",     0.1,    breakdown['isolated_off'],      round(breakdown['isolated_off']*0.1,      2)),
    ("總懲罰值",      "",     "",                             round(raw_penalty, 2)),
]
for r_off, rd in enumerate(bdata):
    r      = br + r_off
    is_hdr = (r_off == 0)
    is_tot = (r_off == len(bdata) - 1)
    bg2    = COLOR_HDR2 if is_hdr else ("FFD700" if is_tot else "FFFFFF")
    fc2    = "FFFFFF" if is_hdr else "000000"
    for co, v in enumerate(rd, 1):
        cs(ws2, r, co, v, bg2, bold=(is_hdr or is_tot), fc=fc2, fs=10)

cws = [14, 10, 10, 12, 14, 12, 12, 12, 10, 10, 12, 14, 14, 14, 14, 14]
for col, w in enumerate(cws, 1):
    ws2.column_dimensions[get_column_letter(col)].width = w

# ─────────────────────────────────────────────
# 8. SAVE
# ─────────────────────────────────────────────

import os
desktop = os.path.join(os.path.expanduser("~"), "Desktop")
os.makedirs(desktop, exist_ok=True)
out_path = os.path.join(desktop, "Scheduling_Output_Gurobi.xlsx")
wb.save(out_path)
print(f"\nExcel saved: {out_path}")
print("All done")


# In[ ]:





import random
import time
from dataclasses import dataclass, asdict
from typing import Dict, List, Tuple
import numpy as np
import pandas as pd
from ortools.sat.python import cp_model

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.comments import Comment
# 演算法（GA、CP-SAT、數學模型）處理數字比處理字串快很多

SHIFT_TO_INT = {"O": 0, "D": 1, "E": 2, "N": 3}
INT_TO_SHIFT = {0: "O", 1: "D", 2: "E", 3: "N"}
WORK_SHIFTS = [1, 2, 3] # 0 是休假，1,2,3是早班、午班、晚班
SHIFT_INDEX = {1: 0, 2: 1, 3: 2}# 需求 = [早班需求, 午班需求, 晚班需求]= [5, 4, 3]

start = time.time() #紀錄時間

@dataclass
class GAConfig:
    population_size: int = 50 # 一次同時嘗試 50 種不同排班表(每一代有 50個排班解)
    generations: int = 500 # 算法最多進化 500代
    tournament_size: int = 4 # 隨機抓4個解 → 選最好的
    crossover_rate: float = 0.90 #隨機抓4個解 → 選最好的
    mutation_rate: float = 0.15 #15% 機率突變
    elite_size: int = 10 #每代保留 最好的10個解
    random_immigrants: int = 6 # 每代加入 6個隨機新解，多樣性維持策略
    local_search_every: int = 8 # 每8代做一次 局部搜尋
    local_search_steps: int = 10 # 每次局部搜尋最多調整 10 次
    stagnation_limit: int = 30 # 如果 30代都沒進步，就會觸發restart 或 immigration 增強
    restarts: int = 5 # 最多允許 15次重啟族群 ，強力跳出 local optimum 手段
    seed: int = 42
    hard_scale: float = 1000.0 # Hard constraint 懲罰倍率

    # 題目規則權重
    w_consecutive_6: float = 1.0
    w_night_to_day_evening: float = 1.0
    w_evening_to_day: float = 1.0
    w_day_evening_to_night: float = 1.0
    w_preset_violation: float = 0.2

    # 題目公平性權重
    w_few_two_day_off_blocks: float = 0.1
    w_monthly_off_lt9: float = 0.1
    w_weekend_off_lt4: float = 0.1
    w_single_day_off: float = 0.1

    # 其他控制參數
    early_stop_rounds: int = 70 #如果 70代都沒有找到更好的解，早停
    local_search_top_k: int = 3 #每次做 Local Search 時，只對前3名最好的排班解做精修
    print_every: int = 20

    # 崔：為甚麼是 0.61 呢呵呵，因為 IEEE 754: 0.2+0.2+0.2 = 0.6000000000000001
    # 嘗試過 0.6 會跑完 15 輪
    target_fitness: float = 0.61 # 若最佳分數 <= 0.61，立即停止
    
    #GA 找到「不錯的解」，再用 CP-SAT 做「數學最佳化精修」
    # CP-SAT 微調參數
    use_cpsat_tune: bool = True #是否開啟 CP-SAT 精修
    cpsat_tune_every: int = 12 #每12代做一次 CP-SAT 微調
    cpsat_tune_trigger: int = 8 #如果 GA 8代沒進步，就提前觸發 CP-SAT
    cpsat_window_days: int = 7# CP-SAT 只優化，7天的排班區段
    cpsat_max_employees: int = 8 #每次 CP-SAT 只選最多6位員工做精修
    cpsat_time_limit: float = 3.0 #CP-SAT 最多算3秒
    cpsat_num_workers: int = 8 #8核心平行搜尋

    # CP-SAT 全域重啟 / 大範圍重整參數
    use_global_cpsat_restart: bool = True
    global_cpsat_trigger: int = 18
    global_cpsat_time_limit: float = 10.0
    global_cpsat_num_workers: int = 8
    global_cpsat_max_change_cells: int = 32# 最多改動 32 個班表格子
    global_cpsat_once_per_run: bool = False #整個演算法可以做多次global repair
    max_global_restarts_per_run: int = 2 #整場 GA 最多做2次全域重整

    # final polishing mode（尾段專打 single-off）
    use_final_polishing: bool = True
    final_polish_threshold: float = 1.0 #如果 penalty 已經低於 1，進入精修
    final_polish_trigger: int = 10
    final_polish_every: int = 4
    final_polish_window_days: int = 14
    final_polish_max_employees: int = 10
    final_polish_time_limit: float = 8.0
    final_polish_num_workers: int = 8
    final_polish_rounds: int = 3 #最終精修最多做3輪
    final_polish_max_change_cells: int = 22 #每輪最多改22個班格


class SchedulerGA:
    # 建立排班系統
    def __init__(self, engineer_csv: str | None = None, demand_csv: str | None = None,
                 config: GAConfig | None = None, *,
                 engineer_df: pd.DataFrame | None = None,
                 demand_df: pd.DataFrame | None = None):
        self.config = config or GAConfig()
        self.rng = random.Random(self.config.seed)
        np.random.seed(self.config.seed)

        if engineer_df is not None and demand_df is not None:
            self.engineer_df = engineer_df.reset_index(drop=True)
            self.demand_df = demand_df.reset_index(drop=True)
        elif engineer_csv is not None and demand_csv is not None:
            self.engineer_df = pd.read_csv(engineer_csv)
            self.demand_df = pd.read_csv(demand_csv)
        else:
            raise ValueError("必須提供 (engineer_csv, demand_csv) 或 (engineer_df, demand_df)")

        self.people_col = "人員"
        self.group_col = "班別群組"
        self.day_cols = [c for c in self.engineer_df.columns if c.startswith("Date_")]

        self.people = self.engineer_df[self.people_col].astype(str).tolist()
        self.groups = [SHIFT_TO_INT.get(str(x).strip().upper()[:1], 0) for x in self.engineer_df[self.group_col].fillna("")]
        self.n_people = len(self.people)
        self.n_days = len(self.day_cols)

        if len(self.demand_df) != self.n_days:
            raise ValueError("Shift_Demand 的天數與 Engineer_List 不一致。")

        self.fixed_mask = np.zeros((self.n_people, self.n_days), dtype=bool)
        self.fixed_values = np.zeros((self.n_people, self.n_days), dtype=np.int8)

        for i in range(self.n_people):
            for d, col in enumerate(self.day_cols):
                val = self.engineer_df.at[i, col]
                if pd.notna(val) and str(val).strip() != "":
                    v = str(val).strip().upper()
                    if v not in SHIFT_TO_INT:
                        raise ValueError(f"未知班別: {v} @ {self.people[i]} {col}")
                    self.fixed_mask[i, d] = True
                    self.fixed_values[i, d] = SHIFT_TO_INT[v]

        self.day_demand = np.zeros((self.n_days, 3), dtype=np.int16)
        self.weekend_flags = np.zeros(self.n_days, dtype=bool)

        for d in range(self.n_days):
            row = self.demand_df.iloc[d]
            self.day_demand[d, 0] = int(row["Day"])
            self.day_demand[d, 1] = int(row["Afternoon"])
            self.day_demand[d, 2] = int(row["Night"])
            wk = row.get("IfWeekend", "")
            if isinstance(wk, bool):
                self.weekend_flags[d] = wk
            else:
                self.weekend_flags[d] = str(wk).strip().upper() in ("Y", "TRUE", "1")

        self._validate_inputs()
    #檢查輸入資料有沒有先天違規
    def _validate_inputs(self) -> None:
        for d in range(self.n_days):
            fixed_counts = np.zeros(3, dtype=int)
            idxs = np.where(self.fixed_mask[:, d])[0]
            for i in idxs:
                s = self.fixed_values[i, d]
                if s in WORK_SHIFTS:
                    fixed_counts[SHIFT_INDEX[s]] += 1
            if np.any(fixed_counts > self.day_demand[d]):
                raise ValueError(f"{self.day_cols[d]} 的預設班別數量超過需求，資料本身不可行。")
    #重設亂數種子
    def reseed(self, seed: int) -> None:
        self.rng = random.Random(seed)
        np.random.seed(seed)
    #動態調整突變率
    def current_mutation_rate(self, generation: int, stagnation: int) -> float:
        base_rate = self.config.mutation_rate
        rate = max(0.04, base_rate * (1.0 - generation / max(1, self.config.generations) * 0.55))
        if stagnation >= self.config.stagnation_limit:
            rate = min(0.32, rate * 2.4)
        elif stagnation >= self.config.stagnation_limit // 2:
            rate = min(0.22, rate * 1.6)
        return rate

    def print_breakdown(self, run_id: int, gen: int, best_breakdown: Dict[str, float], current_mutation: float, stagnation: int) -> None:
        print(
            f"[Run {run_id}][Gen {gen}] "
            f"total={best_breakdown['total_fitness']:.2f} | "
            f"hard={best_breakdown['hard_penalty_raw']:.2f} | "
            f"soft={best_breakdown['soft_penalty_raw']:.2f} | "
            f"mut={current_mutation:.3f} | "
            f"stagnation={stagnation}"
        )
        print(
            "  hard: "
            f"demand={best_breakdown['demand_mismatch']:.2f}, "
            f"preset={best_breakdown['preset_violation']:.2f}, "
            f"consec_6={best_breakdown['consecutive_6']:.2f}, "
            f"trans={best_breakdown['transition_violation']:.2f}"
        )
        print(
            "  soft: "
            f"two_day_off={best_breakdown['few_two_day_off_blocks']:.2f}, "
            f"off_lt9={best_breakdown['monthly_off_lt9']:.2f}, "
            f"weekend_lt4={best_breakdown['weekend_off_lt4']:.2f}, "
            f"single_off={best_breakdown['single_off_days']:.2f}"
        )
    # 先建立空白班表，放入固定班，一天一天排，排完後再修補
    def generate_individual(self) -> np.ndarray:
        schedule = np.zeros((self.n_people, self.n_days), dtype=np.int8)
        schedule[self.fixed_mask] = self.fixed_values[self.fixed_mask]

        for d in range(self.n_days):
            day = schedule[:, d]
            candidates = [i for i in range(self.n_people) if not self.fixed_mask[i, d]]

            for shift in WORK_SHIFTS:
                assigned_now = int(np.sum(day == shift))
                need = self.day_demand[d, SHIFT_INDEX[shift]] - assigned_now
                if need <= 0:
                    continue

                scored = []
                for i in candidates:
                    #如果這個人今天已經有班了，就不能再分第二個班。
                    if day[i] != 0:
                        continue
                    score = 0.0
                    #如果這位員工的預設班別群組剛好就是這個 shift，就加 6 分
                    if self.groups[i] == shift:
                        score += 6.0
                    #如果不是第一天，就去看這個人前一天排什麼班，如果是第一天，前一天視為 0。
                    prev_s = schedule[i, d - 1] if d > 0 else 0
                    #如果會違反轉班規則，扣分
                    if self._transition_violation(prev_s, shift):
                        score -= 12.0
                    # 如果會造成連六，扣分
                    if self._would_hit_streak6(schedule, i, d, shift):
                        score -= 9.0
                    #加一點隨機數，增加多樣性
                    score += self.rng.random()
                    scored.append((score, i))
                #分數高的排前面
                scored.sort(reverse=True)
                chosen = [i for _, i in scored[:need]]
                #把這些人今天的班別設成目前這個 shift。
                for i in chosen:
                    day[i] = shift
            #把更新後的當天班表寫回完整的 schedule
            schedule[:, d] = day
        #先用啟發式方法排出一個「大致合理」的班表，再交給 repair() 去修正缺人、違規、不公平等問題
        schedule = self.repair(schedule, fix_single_off=False)
        return schedule

    def _would_hit_streak6(self, schedule: np.ndarray, i: int, d: int, new_shift: int) -> bool:
        if new_shift not in WORK_SHIFTS:#如果今天是休假，不可能造成連六，所以直接 False
            return False
        streak = 1
        k = d - 1
        while k >= 0 and schedule[i, k] in WORK_SHIFTS:#從前一天開始往回看，只要前一天也是工作班，就連班數 +1，再往前一天看，一直到遇到休假或到最前面為止
            streak += 1
            k -= 1
        return streak >= 6#只要今天排下去之後，連續工作天數 >= 6，就回傳 True。

    def _transition_violation(self, prev_s: int, cur_s: int) -> bool:
        if prev_s == 3 and cur_s in {1, 2}:#前一天班別 prev_s 到今天班別 cur_s，是否屬於定義的違規轉班
            return True
        if prev_s == 2 and cur_s == 1:
            return True
        if prev_s in {1, 2} and cur_s == 3:
            return True
        return False

    def repair(self, schedule: np.ndarray, fix_single_off: bool = False) -> np.ndarray:
        s = schedule.copy()
        # 1. 強制刷回固定值 (確保比賽指定的固定格不可動)
        s[self.fixed_mask] = self.fixed_values[self.fixed_mask]

        # 2. 初步修復每日人數需求 (確保 hard: demand = 0)
        for d in range(self.n_days):
            self._repair_day_counts(s, d)

        # 3. 核心法規修復 (循環處理轉換與連6限制)
        for _ in range(3): # 減少循環次數以提升速度
            changed = False
            for i in range(self.n_people):
                # 這裡應優先修復連 6 與非法班別銜接 (權重 1.0)
                if self._repair_person_transitions(s, i):
                    changed = True
                if self._repair_person_streaks(s, i):
                    changed = True
            if not changed: break

        # 4. 針對性優化：消除單休 (Soft Penalty)
        # 改成可選，不在 GA 前段每次都做，避免過早收斂
        if fix_single_off:
            self._repair_single_off_targeted(s)

        # 5. 最終人數校正與固定值保護 (Double Check)
        for d in range(self.n_days):
            self._repair_day_counts(s, d)
        
        s[self.fixed_mask] = self.fixed_values[self.fixed_mask]
        return s

    def _repair_single_off_targeted(self, schedule: np.ndarray) -> None:
        """
        在不增加『違反班表』與『連6』的前提下，消除單休模式。
        """
        for i in range(self.n_people):
            default_i = self.groups[i]  # 取得員工 i 的預設班別群組 [cite: 54]
            for d in range(1, self.n_days - 1):
                # 偵測到單休 [工作-休-工作] [cite: 28, 529, 694]
                if schedule[i, d] == 0 and schedule[i, d-1] != 0 and schedule[i, d+1] != 0:
                    if self.fixed_mask[i, d]: continue
                    
                    # 優先尋找群組一致且補班後安全的人進行交換
                    # 條件：j 當天排班與 i 的預設群組相同，且 j 換成休假後也符合其預設 (或是換成休假不罰)
                    candidates = [
                        j for j in range(self.n_people)
                        if not self.fixed_mask[j, d] 
                        and schedule[j, d] == default_i      # 關鍵：j 讓出的班別剛好是 i 想要的 [cite: 51, 294]
                        and self._check_consecutive_safe(schedule, i, d) # 確保 i 不會連 6 [cite: 18, 160, 222]
                    ]
                    
                    if not candidates:
                        # 如果找不到完全符合群組的人，才考慮其他不會增加 preset 懲罰的人 (例如 j 原本就在違規的人)
                        candidates = [
                            j for j in range(self.n_people)
                            if not self.fixed_mask[j, d] and schedule[j, d] != 0
                            and self._check_consecutive_safe(schedule, i, d)
                        ]
                    
                    if candidates:
                        j = self.rng.choice(candidates)
                        # 執行交換
                        schedule[i, d], schedule[j, d] = schedule[j, d], schedule[i, d]

    def _check_consecutive_safe(self, schedule, i, d) -> bool:
        """檢查員工 i 在第 d 天上班後，總連續上班天數是否 <= 5 (預留 1 天緩衝避免罰分)"""
        left = 0
        for day in range(d - 1, -1, -1):
            if schedule[i, day] != 0: left += 1
            else: break
        right = 0
        for day in range(d + 1, self.n_days):
            if schedule[i, day] != 0: right += 1
            else: break
        # 若總長度 > 5 則回傳 False (不執行交換)
        return (left + 1 + right) <= 5
    # 把這一天的班表修到剛好合法又比較公平
    def _repair_day_counts(self, schedule: np.ndarray, d: int) -> None:
        day = schedule[:, d]
        counts = np.array([np.sum(day == 1), np.sum(day == 2), np.sum(day == 3)], dtype=int)
        is_weekend = self.weekend_flags[d]

        for shift in WORK_SHIFTS:
            idx = SHIFT_INDEX[shift]
            overflow = int(counts[idx] - self.day_demand[d, idx])
            if overflow <= 0: continue

            # --- 當人手過多時，優先讓誰去休假 (解決 weekend_lt4 與 single_off) ---
            candidates = [i for i in range(self.n_people) if day[i] == shift and not self.fixed_mask[i, d]]
            scored = []
            for i in candidates:
                score = 0.0
                # 1. 如果今天是週末，且該員週末休假 < 4，給予極高權限去休假
                if is_weekend:
                    weekend_off = np.sum((schedule[i] == 0) & self.weekend_flags)
                    if weekend_off < 4: score += 30.0
                
                # 2. 如果這格是「單休 1 日」的潛在位置，優先砍掉變成「連休」
                if d > 0 and d < self.n_days - 1:
                    if schedule[i, d-1] == 0 or schedule[i, d+1] == 0:
                        score += 15.0
                
                score += self.rng.random()
                scored.append((score, i))

            scored.sort(reverse=True)
            for _, i in scored[:overflow]:
                day[i] = 0
            counts[idx] -= overflow

        for shift in WORK_SHIFTS:
            idx = SHIFT_INDEX[shift]
            deficit = int(self.day_demand[d, idx] - counts[idx])
            if deficit <= 0: continue

            # --- 當缺人手時，優先叫誰回來上班 (避開正在休假的人) ---
            candidates = [i for i in range(self.n_people) if day[i] == 0 and not self.fixed_mask[i, d]]
            scored = []
            for i in candidates:
                score = 0.0
                # 1. 符合預設班別群組的人優先回來
                if self.groups[i] == shift: score += 10.0
                
                # 2. 如果今天是週末，且該員週末休假已經 >= 4，叫他回來上班，把休假留給別人
                if is_weekend:
                    weekend_off = np.sum((schedule[i] == 0) & self.weekend_flags)
                    if weekend_off >= 4: score += 5.0
                
                # 3. 避免破壞別人的「連休」 (如果前後都是休假，就不要叫他回來)
                if d > 0 and d < self.n_days - 1:
                    if schedule[i, d-1] == 0 and schedule[i, d+1] == 0:
                        score -= 15.0
                
                # 檢查法律限制 (絕對不能違反)
                prev_s = schedule[i, d - 1] if d > 0 else 0
                if self._transition_violation(prev_s, shift): score -= 100.0
                if self._would_hit_streak6(schedule, i, d, shift): score -= 100.0

                score += self.rng.random()
                scored.append((score, i))

            scored.sort(reverse=True)
            for _, i in scored[:deficit]:
                day[i] = shift
            counts[idx] += deficit

        schedule[:, d] = day
    # 某位員工在某一天是否屬於連續工作 ≥6 天的違規區段
    # def _is_streak_violation_at(self, schedule: np.ndarray, i: int, d: int) -> bool:
    #     if schedule[i, d] not in WORK_SHIFTS:
    #         return False
    #     left = d
    #     while left - 1 >= 0 and schedule[i, left - 1] in WORK_SHIFTS:
    #         left -= 1
    #     right = d
    #     while right + 1 < self.n_days and schedule[i, right + 1] in WORK_SHIFTS:
    #         right += 1
    #     return (right - left + 1) >= 6
    # 修個人班別轉換違規
    def _repair_person_transitions(self, schedule: np.ndarray, i: int) -> bool:
        changed = False
        for d in range(1, self.n_days):
            prev_s = schedule[i, d - 1]
            cur_s = schedule[i, d]
            if not self._transition_violation(prev_s, cur_s):
                continue

            if not self.fixed_mask[i, d]:
                if cur_s in WORK_SHIFTS and np.sum(schedule[:, d] == cur_s) - 1 >= self.day_demand[d, SHIFT_INDEX[cur_s]]:
                    schedule[i, d] = 0
                    changed = True
                    continue

                for alt in [1, 2, 3, 0]:
                    if alt == cur_s:
                        continue
                    if self._transition_violation(prev_s, alt):
                        continue
                    if alt in WORK_SHIFTS and np.sum(schedule[:, d] == alt) + 1 > self.day_demand[d, SHIFT_INDEX[alt]]:
                        continue
                    if cur_s in WORK_SHIFTS and np.sum(schedule[:, d] == cur_s) - 1 < self.day_demand[d, SHIFT_INDEX[cur_s]]:
                        continue
                    schedule[i, d] = alt
                    changed = True
                    break

            elif not self.fixed_mask[i, d - 1]:
                if prev_s in WORK_SHIFTS and np.sum(schedule[:, d - 1] == prev_s) - 1 >= self.day_demand[d - 1, SHIFT_INDEX[prev_s]]:
                    schedule[i, d - 1] = 0
                    changed = True
        return changed
    #某位員工在某一天是否屬於連續工作 ≥6 天的違規區段
    def _repair_person_streaks(self, schedule: np.ndarray, i: int) -> bool:
        changed = False
        streak = 0
        for d in range(self.n_days):
            if schedule[i, d] in WORK_SHIFTS:
                streak += 1
            else:
                streak = 0

            if streak >= 6:
                if (not self.fixed_mask[i, d]) and (np.sum(schedule[:, d] == schedule[i, d]) - 1 >= self.day_demand[d, SHIFT_INDEX[schedule[i, d]]]):
                    schedule[i, d] = 0
                    changed = True
                    streak = 0
                    continue

                start = d
                while start - 1 >= 0 and schedule[i, start - 1] in WORK_SHIFTS:
                    start -= 1
                repaired = False
                for x in range(d, start - 1, -1):
                    if self.fixed_mask[i, x]:
                        continue
                    cur_s = schedule[i, x]
                    if cur_s in WORK_SHIFTS and np.sum(schedule[:, x] == cur_s) - 1 >= self.day_demand[x, SHIFT_INDEX[cur_s]]:
                        schedule[i, x] = 0
                        changed = True
                        streak = 0
                        repaired = True
                        break

                if not repaired:
                    cur_s = schedule[i, d]
                    off_candidates = [j for j in range(self.n_people) if j != i and schedule[j, d] == 0 and not self.fixed_mask[j, d]]
                    self.rng.shuffle(off_candidates)
                    for j in off_candidates:
                        prev_j = schedule[j, d - 1] if d > 0 else 0
                        if not self._transition_violation(prev_j, cur_s):
                            schedule[j, d] = cur_s
                            schedule[i, d] = 0
                            changed = True
                            streak = 0
                            break
        return changed

    # def _repair_fairness(self, schedule: np.ndarray) -> None:
    #     for _ in range(2):
    #         for i in range(self.n_people):
    #             for d in range(1, self.n_days - 1):
    #                 if schedule[i, d] == 0 and schedule[i, d - 1] in WORK_SHIFTS and schedule[i, d + 1] in WORK_SHIFTS:
    #                     if (not self.fixed_mask[i, d - 1]) and (np.sum(schedule[:, d - 1] == schedule[i, d - 1]) - 1 >= self.day_demand[d - 1, SHIFT_INDEX[schedule[i, d - 1]]]):
    #                         schedule[i, d - 1] = 0
    #                     elif (not self.fixed_mask[i, d + 1]) and (np.sum(schedule[:, d + 1] == schedule[i, d + 1]) - 1 >= self.day_demand[d + 1, SHIFT_INDEX[schedule[i, d + 1]]]):
    #                         schedule[i, d + 1] = 0
    #         for d in range(self.n_days):
    #             self._repair_day_counts(schedule, d)

    #計算這張排班表有多好（或多爛）
    def fitness(self, schedule: np.ndarray) -> Tuple[float, Dict[str, float]]:
        breakdown: Dict[str, float] = {}
        hard_penalty = 0.0  # 法規與硬性約束
        soft_penalty = 0.0  # 公平性約束

        # 1. 班別需求滿足度 (滿足每日各班別人力需求) 
        demand_mismatch = 0
        for d in range(self.n_days):
            demand_mismatch += abs(int(np.sum(schedule[:, d] == 1)) - int(self.day_demand[d, 0]))
            demand_mismatch += abs(int(np.sum(schedule[:, d] == 2)) - int(self.day_demand[d, 1]))
            demand_mismatch += abs(int(np.sum(schedule[:, d] == 3)) - int(self.day_demand[d, 2]))
        breakdown["demand_mismatch"] = float(demand_mismatch)
        hard_penalty += float(demand_mismatch)

        # 2. 違反預設班別 (權重 0.2) 
        # 規則：人員被分配到「非休假(O)」且「非預設班別群組」時計入懲罰 [cite: 51]
        preset_viol_count = 0
        for i in range(self.n_people):
            default_s = self.groups[i]  # 從「班別群組」欄位讀取的代號 (1:D, 2:E, 3:N)
            for d in range(self.n_days):
                actual_s = schedule[i, d]
                # 判定：如果這格不是休假(0)，且不等於他的預設群組，就屬違規
                if actual_s != 0 and actual_s != default_s:
                    preset_viol_count += 1
        
        breakdown["preset_violation"] = preset_viol_count * self.config.w_preset_violation
        # 將此懲罰計入 hard_penalty 或目標總分 [cite: 13, 17]
        hard_penalty += breakdown["preset_violation"]

        consec6_count = 0
        n2de_count = 0  # 夜接早/午 
        e2d_count = 0   # 午接早 
        de2n_count = 0  # 早/午接夜 

        few_two_day_off_blocks = 0
        off_lt9_penalty = 0
        weekend_off_lt4_penalty = 0
        single_off_count = 0

        for i in range(self.n_people):
            row = schedule[i, :]

            # 3. 連續上班 6 天 (權重 1.0) 
            # 規則：連續 6 天罰 1 次，連續 7 天罰 2 次，以此類推 
            streak = 0
            for d in range(self.n_days):
                if row[d] in WORK_SHIFTS:
                    streak += 1
                    if streak >= 6:
                        consec6_count += 1
                else:
                    streak = 0

            # 4. 班別轉換限制 (權重 1.0) 
            for d in range(1, self.n_days):
                prev_s, cur_s = row[d - 1], row[d]
                if prev_s == 3 and cur_s in {1, 2}:  # 夜不可接早/午 
                    n2de_count += 1
                if prev_s == 2 and cur_s == 1:       # 午不可接早 
                    e2d_count += 1
                if prev_s in {1, 2} and cur_s == 3:  # 早/午不可接夜 
                    de2n_count += 1

            # 5. 每人每月連續休假數小於 2 次 (權重 0.1) 
            # 規則：連續休假 >= 2 天才算一次；不足 2 次則罰 0.1 
            d = 0
            off_blocks_ge2 = 0
            while d < self.n_days:
                if row[d] == 0:
                    start = d
                    while d < self.n_days and row[d] == 0:
                        d += 1
                    if (d - start) >= 2:
                        off_blocks_ge2 += 1
                else:
                    d += 1
            if off_blocks_ge2 < 2:
                few_two_day_off_blocks += 1

            # 6. 每人每月休假天數小於 9 天 (權重 0.1) 
            # 規則：少 1 天罰 0.1，少 2 天罰 0.2，以此類推 
            off_count = int(np.sum(row == 0))
            if off_count < 9:
                off_lt9_penalty += (9 - off_count)

            # 7. 每人每月週末休假天數小於 4 天 (權重 0.1)
            # 規則：少 1 天罰 0.1，以此類推 [cite: 608]
            weekend_off = int(np.sum((row == 0) & self.weekend_flags))
            if weekend_off < 4:
                weekend_off_lt4_penalty += (4 - weekend_off)

            # 8. 僅排休 1 日 (非連續休假) (權重 0.1)
            # 規則：出現「工作+休假+工作」罰 1 次 
            # 註：第一天與最後一日之單獨休假不計懲罰 
            for d in range(1, self.n_days - 1):
                if row[d] == 0 and row[d-1] in WORK_SHIFTS and row[d+1] in WORK_SHIFTS:
                    single_off_count += 1

        # 加總所有權重 [cite: 164, 231, 297]
        hard_penalty += consec6_count * self.config.w_consecutive_6
        hard_penalty += (n2de_count + e2d_count + de2n_count) * 1.0 # 班別轉換權重皆為 1.0

        soft_penalty += few_two_day_off_blocks * self.config.w_few_two_day_off_blocks
        soft_penalty += off_lt9_penalty * self.config.w_monthly_off_lt9
        soft_penalty += weekend_off_lt4_penalty * self.config.w_weekend_off_lt4
        soft_penalty += single_off_count * self.config.w_single_day_off

        # 整理輸出細目
        breakdown["consecutive_6"] = consec6_count * self.config.w_consecutive_6
        breakdown["transition_violation"] = (n2de_count + e2d_count + de2n_count) * 1.0
        breakdown["few_two_day_off_blocks"] = few_two_day_off_blocks * self.config.w_few_two_day_off_blocks
        breakdown["monthly_off_lt9"] = off_lt9_penalty * self.config.w_monthly_off_lt9
        breakdown["weekend_off_lt4"] = weekend_off_lt4_penalty * self.config.w_weekend_off_lt4
        breakdown["single_off_days"] = single_off_count * self.config.w_single_day_off
        
        breakdown["hard_penalty_raw"] = hard_penalty
        breakdown["soft_penalty_raw"] = soft_penalty
        
        # 演算法內部仍用 1000 倍率計算
        weighted_total = (self.config.hard_scale * hard_penalty) + soft_penalty
        # 但在 breakdown 中記錄原始加總 
        breakdown["total_fitness"] = hard_penalty + soft_penalty 
        
        return weighted_total, breakdown
    # 把排班表中所有違規的位置與原因『整理出來』，方便標註、輸出報表或 debug
    def collect_violations(self, schedule: np.ndarray) -> Dict[str, List]:
        violations = {
            "preset_violation_cells": [],
            "transition_cells": [],
            "consecutive_6_cells": [],
            "single_off_cells": [],
            "employee_flags": [],
        }

        
        for i in range(self.n_people):
            default_s = self.groups[i]
            for d in range(self.n_days):
                actual_s = schedule[i, d]
                if actual_s != 0 and actual_s != default_s:
                    msg = f"違反預設班別：預設為 {INT_TO_SHIFT.get(default_s, 'O')}，實際為 {INT_TO_SHIFT[int(actual_s)]}"
                    violations["preset_violation_cells"].append((i, d, msg))

        for i in range(self.n_people):
            row = schedule[i, :]
            for d in range(1, self.n_days):
                prev_s, cur_s = row[d - 1], row[d]
                if prev_s == 3 and cur_s in {1, 2}:
                    violations["transition_cells"].append((i, d, "前一天夜班不可接早班/午班"))
                if prev_s == 2 and cur_s == 1:
                    violations["transition_cells"].append((i, d, "前一天午班不可接早班"))
                if prev_s in {1, 2} and cur_s == 3:
                    violations["transition_cells"].append((i, d, "前一天早班/午班不可接夜班"))

        for i in range(self.n_people):
            streak = 0
            for d in range(self.n_days):
                if schedule[i, d] in WORK_SHIFTS:
                    streak += 1
                    if streak >= 6:
                        violations["consecutive_6_cells"].append((i, d, "連續上班6天"))
                else:
                    streak = 0

        for i in range(self.n_people):
            for d in range(1, self.n_days - 1):
                if schedule[i, d] == 0 and schedule[i, d - 1] in WORK_SHIFTS and schedule[i, d + 1] in WORK_SHIFTS:
                    violations["single_off_cells"].append((i, d, "僅排休1日(非連續休假)"))

        for i in range(self.n_people):
            row = schedule[i, :]
            off_count = int(np.sum(row == 0))
            weekend_off = int(np.sum((row == 0) & self.weekend_flags))

            d = 0
            off_blocks_ge2 = 0
            while d < self.n_days:
                if row[d] == 0:
                    start = d
                    while d < self.n_days and row[d] == 0:
                        d += 1
                    if d - start >= 2:
                        off_blocks_ge2 += 1
                else:
                    d += 1

            reasons = []
            if off_blocks_ge2 < 2:
                reasons.append(f"連續休假(>=2天)次數不足：{off_blocks_ge2}")
            if off_count < 9:
                reasons.append(f"每月休假少於9天：{off_count}")
            if weekend_off < 4:
                reasons.append(f"週末休假少於4天：{weekend_off}")

            if reasons:
                violations["employee_flags"].append((i, reasons))

        return violations
    # 計算整個族群每張排班表的 fitness 分數
    def evaluate_population(self, population: List[np.ndarray]) -> List[float]:
        return [self.fitness(ind)[0] for ind in population]
    #用「錦標賽選擇法」選父母，隨機抽 k 個個體，選 fitness 最好的當作父母
    def tournament_select(self, population: List[np.ndarray], fitnesses: List[float]) -> np.ndarray:
        idxs = self.rng.sample(range(len(population)), self.config.tournament_size)
        best_idx = min(idxs, key=lambda idx: fitnesses[idx])
        return population[best_idx].copy()
    #讓兩張排班表「混血」交配
    def crossover(self, p1: np.ndarray, p2: np.ndarray) -> Tuple[np.ndarray, np.ndarray]:
        c1, c2 = p1.copy(), p2.copy()
        if self.rng.random() > self.config.crossover_rate:
            return c1, c2

        row_mask = np.random.rand(self.n_people) < 0.5#人員層 crossover
        c1[row_mask, :] = p2[row_mask, :]#時間區段 crossover
        c2[row_mask, :] = p1[row_mask, :]

        if self.n_days >= 4:
            a = self.rng.randint(0, self.n_days - 2)
            b = self.rng.randint(a + 1, self.n_days - 1)
            c1[:, a:b] = p2[:, a:b]
            c2[:, a:b] = p1[:, a:b]

        c1 = self.repair(c1, fix_single_off=False)#crossover 很容易違法，repair 保證解仍可行
        c2 = self.repair(c2, fix_single_off=False)
        return c1, c2
    #隨機改動排班，增加多樣性
    def mutate(self, schedule: np.ndarray, generation: int, stagnation: int) -> np.ndarray:
        child = schedule.copy()
        decay_rate = self.current_mutation_rate(generation, stagnation)
        if self.rng.random() > decay_rate:
            return child

        n_moves = 1 if stagnation < self.config.stagnation_limit // 2 else 2
        if stagnation >= self.config.stagnation_limit:
            n_moves = 3

        for _ in range(n_moves):
            move_type = self.rng.choice(["swap_same_day", "reassign_one", "swap_blocks", "targeted_fix"])

            if move_type == "swap_same_day":#同一天兩人互換班
                d = self.rng.randrange(self.n_days)
                i, j = self.rng.sample(range(self.n_people), 2)
                if (not self.fixed_mask[i, d]) and (not self.fixed_mask[j, d]):
                    child[i, d], child[j, d] = child[j, d], child[i, d]

            elif move_type == "reassign_one":#某人某天改成任一班
                d = self.rng.randrange(self.n_days)
                i = self.rng.randrange(self.n_people)
                if not self.fixed_mask[i, d]:
                    child[i, d] = self.rng.choice([0, 1, 2, 3])

            elif move_type == "swap_blocks":#兩人交換連續幾天班表
                i, j = self.rng.sample(range(self.n_people), 2)
                start = self.rng.randrange(self.n_days)
                length = self.rng.randint(2, 4)
                end = min(self.n_days, start + length)
                for d in range(start, end):
                    if (not self.fixed_mask[i, d]) and (not self.fixed_mask[j, d]):
                        child[i, d], child[j, d] = child[j, d], child[i, d]

            else:# 如果發現違規轉班，直接讓那天休假
                i = self.rng.randrange(self.n_people)
                for d in range(1, self.n_days):
                    if self._transition_violation(child[i, d - 1], child[i, d]) and (not self.fixed_mask[i, d]):
                        child[i, d] = 0
                        break

        child = self.repair(child, fix_single_off=False)#mutation 很可能破壞合法性，要repair
        return child
    #在 GA 找到一個解後，用「小幅交換」嘗試讓解變更好。
    def local_search(self, schedule: np.ndarray) -> np.ndarray:
        best = schedule.copy()
        best_fit, _ = self.fitness(best)
        for _ in range(self.config.local_search_steps):
            cand = best.copy()
            d = self.rng.randrange(self.n_days)
            i, j = self.rng.sample(range(self.n_people), 2)
            if self.fixed_mask[i, d] or self.fixed_mask[j, d]:
                continue
            cand[i, d], cand[j, d] = cand[j, d], cand[i, d]
            cand = self.repair(cand, fix_single_off=False)
            cand_fit, _ = self.fitness(cand)
            if cand_fit < best_fit:
                best, best_fit = cand, cand_fit
        return best

    # 評估某位員工的排班有多糟，用來決定 CP-SAT 要優先修誰
    def _employee_problem_score(self, schedule: np.ndarray, i: int) -> float:
        row = schedule[i, :]
        single = 0
        for d in range(1, self.n_days - 1):
            if row[d] == 0 and row[d - 1] in WORK_SHIFTS and row[d + 1] in WORK_SHIFTS:
                single += 1
        weekend_off = int(np.sum((row == 0) & self.weekend_flags))
        weekend_short = max(0, 4 - weekend_off)
        preset = int(np.sum((row != 0) & (row != self.groups[i])))
        return single * 3.0 + weekend_short * 2.0 + preset * 0.5
    # 找出「問題最嚴重的時間區段」，只在那幾天做數學最佳化，分數最高的天當中心
    def _pick_cpsat_window(self, schedule: np.ndarray, focus_employees: List[int]) -> List[int]:
        scores = np.zeros(self.n_days, dtype=float)
        for i in focus_employees:
            row = schedule[i, :]
            for d in range(1, self.n_days - 1):
                if row[d] == 0 and row[d - 1] in WORK_SHIFTS and row[d + 1] in WORK_SHIFTS:
                    scores[d] += 3.0
            weekend_off = int(np.sum((row == 0) & self.weekend_flags))
            if weekend_off < 4:
                for d in range(self.n_days):
                    if self.weekend_flags[d] and row[d] != 0:
                        scores[d] += 2.0
        if float(scores.max()) <= 0:
            center = self.n_days // 2
        else:
            center = int(np.argmax(scores))
        half = max(1, self.config.cpsat_window_days // 2)
        start = max(0, center - half)
        end = min(self.n_days, start + self.config.cpsat_window_days)
        start = max(0, end - self.config.cpsat_window_days)
        return list(range(start, end))
    #回傳某員工某天「是否在工作」，如果不在 → 用原本排班結果
    def _work_expr(self, xmap, base: np.ndarray, i: int, d: int):
        if d < 0 or d >= self.n_days:
            return 0
        key = (i, d)
        if key in xmap:
            return sum(xmap[key][s] for s in WORK_SHIFTS)
        return 1 if base[i, d] in WORK_SHIFTS else 0
    #回傳某員工某天是否為特定班別
    def _shift_expr(self, xmap, base: np.ndarray, i: int, d: int, shift: int):
        if d < 0 or d >= self.n_days:
            return 0
        key = (i, d)
        if key in xmap:
            return xmap[key][shift]
        return 1 if int(base[i, d]) == shift else 0
    # 把 GA 找到的排班，挑一小群員工 + 一小段日期，用 CP-SAT 做精確最佳化修正，局部數學求最佳
    def _solve_cpsat_subproblem(
        self,
        base: np.ndarray,
        employees: List[int],
        days: List[int],
        *,
        time_limit: float,
        num_workers: int,
        max_change_cells: int | None = None,
    ) -> np.ndarray | None:
        model = cp_model.CpModel()
        day_set = set(days)
        emp_set = set(employees)

        xmap = {}
        for i in employees:
            for d in days:
                vars_s = {}
                for s in [0, 1, 2, 3]:
                    vars_s[s] = model.NewBoolVar(f"x_{i}_{d}_{s}")
                model.Add(sum(vars_s.values()) == 1)

                if self.fixed_mask[i, d]:
                    fixed_s = int(self.fixed_values[i, d])
                    model.Add(vars_s[fixed_s] == 1)
                xmap[(i, d)] = vars_s

        # 每日需求精確滿足
        for d in days:
            for shift in WORK_SHIFTS:
                fixed_count = 0
                for i in range(self.n_people):
                    if i in emp_set:
                        continue
                    if int(base[i, d]) == shift:
                        fixed_count += 1
                residual = int(self.day_demand[d, SHIFT_INDEX[shift]] - fixed_count)
                if residual < 0:
                    return None
                model.Add(sum(xmap[(i, d)][shift] for i in employees) == residual)

        # 相鄰班別銜接限制
        forbidden_pairs = [(3, 1), (3, 2), (2, 1), (1, 3), (2, 3)]
        for i in employees:
            for d in range(1, self.n_days):
                if (d not in day_set) and ((d - 1) not in day_set):
                    continue
                for a, b in forbidden_pairs:
                    model.Add(self._shift_expr(xmap, base, i, d - 1, a) + self._shift_expr(xmap, base, i, d, b) <= 1)

        # 任何連續 6 天皆不可全上班
        for i in employees:
            for start in range(0, self.n_days - 5):
                window_days = list(range(start, start + 6))
                if not any(d in day_set for d in window_days):
                    continue
                model.Add(sum(self._work_expr(xmap, base, i, d) for d in window_days) <= 5)

        # 變動格數上限（避免全域大洗牌，保留 GA 找到的主結構）
        if max_change_cells is not None:
            change_vars = []
            for i in employees:
                for d in days:
                    cv = model.NewBoolVar(f"chg_{i}_{d}")
                    model.Add(cv == sum(xmap[(i, d)][s] for s in [0, 1, 2, 3] if s != int(base[i, d])))
                    change_vars.append(cv)
            model.Add(sum(change_vars) <= int(max_change_cells))

        # 目標：壓 preset / single off / weekend shortage / 月休不足
        obj_terms = []

        for i in employees:
            default_s = int(self.groups[i])
            for d in days:
                if default_s != 0:
                    v = model.NewBoolVar(f"preset_{i}_{d}")
                    model.Add(v == sum(xmap[(i, d)][s] for s in WORK_SHIFTS if s != default_s))
                    obj_terms.append(2 * v)

        for i in employees:
            for d in range(1, self.n_days - 1):
                if not (d in day_set or (d - 1) in day_set or (d + 1) in day_set):
                    continue
                off_d = 1 - self._work_expr(xmap, base, i, d)
                prev_w = self._work_expr(xmap, base, i, d - 1)
                next_w = self._work_expr(xmap, base, i, d + 1)
                v = model.NewBoolVar(f"single_{i}_{d}")
                model.Add(v <= off_d)
                model.Add(v <= prev_w)
                model.Add(v <= next_w)
                model.Add(v >= off_d + prev_w + next_w - 2)
                obj_terms.append(1 * v)

            total_weekend_off_outside = int(np.sum((base[i] == 0) & self.weekend_flags))
            inside_weekend_days = [d for d in days if self.weekend_flags[d]]
            if inside_weekend_days:
                outside_count = total_weekend_off_outside - int(np.sum((base[i, inside_weekend_days] == 0)))
                inside_off = sum(xmap[(i, d)][0] for d in inside_weekend_days)
                shortage = model.NewIntVar(0, 4, f"wshort_{i}")
                model.Add(shortage >= 4 - (outside_count + inside_off))
                obj_terms.append(shortage)

            total_off_outside = int(np.sum(base[i] == 0))
            outside_count = total_off_outside - int(np.sum(base[i, days] == 0))
            inside_off = sum(xmap[(i, d)][0] for d in days)
            month_short = model.NewIntVar(0, 9, f"mshort_{i}")
            model.Add(month_short >= 9 - (outside_count + inside_off))
            obj_terms.append(month_short)

        model.Minimize(sum(obj_terms))
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = float(time_limit)
        solver.parameters.num_search_workers = int(num_workers)
        solver.parameters.random_seed = int(self.config.seed)
        status = solver.Solve(model)
        if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
            return None

        tuned = base.copy()
        for i in employees:
            for d in days:
                for s in [0, 1, 2, 3]:
                    if solver.Value(xmap[(i, d)][s]) == 1:
                        tuned[i, d] = s
                        break
        tuned = self.repair(tuned)
        return tuned
    # 用「微調模式」呼叫 CP-SAT 修某些員工 + 某幾天
    def _solve_cpsat_window(self, base: np.ndarray, employees: List[int], days: List[int]) -> np.ndarray | None:
        return self._solve_cpsat_subproblem(
            base,
            employees,
            days,
            time_limit=self.config.cpsat_time_limit,
            num_workers=self.config.cpsat_num_workers,
            max_change_cells=None,
        )
    # 用 CP-SAT 嘗試「全域大重整」
    def cpsat_global_restart(self, schedule: np.ndarray) -> np.ndarray:
        employees = list(range(self.n_people))
        days = list(range(self.n_days))
        tuned = self._solve_cpsat_subproblem(
            schedule,
            employees,
            days,
            time_limit=self.config.global_cpsat_time_limit,
            num_workers=self.config.global_cpsat_num_workers,
            max_change_cells=self.config.global_cpsat_max_change_cells,
        )
        if tuned is None:
            return schedule
        old_fit, _ = self.fitness(schedule)
        new_fit, _ = self.fitness(tuned)
        return tuned if new_fit < old_fit else schedule
    # 找最有問題的幾位員工 → 修他們的問題區段
    def cpsat_micro_tune(self, schedule: np.ndarray) -> np.ndarray:
        ranked = sorted(range(self.n_people), key=lambda i: self._employee_problem_score(schedule, i), reverse=True)
        employees = [i for i in ranked if self._employee_problem_score(schedule, i) > 0][: self.config.cpsat_max_employees]
        if not employees:
            return schedule
        days = self._pick_cpsat_window(schedule, employees)
        tuned = self._solve_cpsat_window(schedule, employees, days)
        if tuned is None:
            return schedule
        old_fit, _ = self.fitness(schedule)
        new_fit, _ = self.fitness(tuned)
        return tuned if new_fit < old_fit else schedule
    # 找某員工所有 single-off 的位置
    def _single_off_positions(self, schedule: np.ndarray, i: int) -> List[int]:
        row = schedule[i, :]
        pos: List[int] = []
        for d in range(1, self.n_days - 1):
            if row[d] == 0 and row[d - 1] in WORK_SHIFTS and row[d + 1] in WORK_SHIFTS:
                pos.append(d)
        return pos
    # 找「single-off 密集區域」做精修
    def _pick_single_off_window(self, schedule: np.ndarray, employees: List[int]) -> List[int]:
        scores = np.zeros(self.n_days, dtype=float)
        for i in employees:
            for d in self._single_off_positions(schedule, i):
                left = max(0, d - 2)
                right = min(self.n_days, d + 3)
                scores[left:right] += 3.0
                if self.weekend_flags[d]:
                    scores[d] += 2.0
            weekend_off = int(np.sum((schedule[i] == 0) & self.weekend_flags))
            if weekend_off < 4:
                for d in range(self.n_days):
                    if self.weekend_flags[d] and schedule[i, d] != 0:
                        left = max(0, d - 4)
                        right = min(self.n_days, d + 5)
                        scores[left:right] += 2.5
        if float(scores.max()) <= 0:
            weekend_days = [d for d in range(self.n_days) if self.weekend_flags[d]]
            center = weekend_days[len(weekend_days) // 2] if weekend_days else self.n_days // 2
        else:
            center = int(np.argmax(scores))
        half = max(1, self.config.final_polish_window_days // 2)
        start = max(0, center - half)
        end = min(self.n_days, start + self.config.final_polish_window_days)
        start = max(0, end - self.config.final_polish_window_days)
        return list(range(start, end))
    # 排班最後的精緻打磨模式，消除 single-off，改善週末休假，不破壞 hard constraint
    def cpsat_final_polish(self, schedule: np.ndarray) -> np.ndarray:
        best = schedule.copy()
        best_fit, best_breakdown = self.fitness(best)
        if best_breakdown["single_off_days"] <= 0 and best_breakdown["weekend_off_lt4"] <= 0:
            return best

        offenders = []
        support = []
        for i in range(self.n_people):
            cnt = len(self._single_off_positions(best, i))
            weekend_off = int(np.sum((best[i] == 0) & self.weekend_flags))
            weekend_short = max(0, 4 - weekend_off)
            total_off = int(np.sum(best[i] == 0))
            if cnt > 0 or weekend_short > 0:
                offenders.append((cnt * 3 + weekend_short * 2, cnt, weekend_short, i))
            else:
                slack = max(0, weekend_off - 4) + 0.5 * max(0, total_off - 9)
                if slack > 0:
                    support.append((slack, weekend_off, total_off, i))
        if not offenders:
            return best

        offenders.sort(reverse=True)
        core_n = min(max(4, self.config.final_polish_max_employees // 2), len(offenders))
        focus = [i for _, _, _, i in offenders[:core_n]]
        focus_set = set(focus)

        support.sort(reverse=True)
        need_support = max(0, self.config.final_polish_max_employees - len(focus))
        for _, _, _, i in support[:need_support]:
            if i not in focus_set:
                focus.append(i)
                focus_set.add(i)

        for _, _, _, i in offenders[core_n:]:
            if len(focus) >= self.config.final_polish_max_employees:
                break
            if i not in focus_set:
                focus.append(i)
                focus_set.add(i)

        for _ in range(self.config.final_polish_rounds):
            days = self._pick_single_off_window(best, focus)
            tuned = self._solve_cpsat_subproblem(
                best,
                focus,
                days,
                time_limit=self.config.final_polish_time_limit,
                num_workers=self.config.final_polish_num_workers,
                max_change_cells=self.config.final_polish_max_change_cells,
            )
            if tuned is None:
                break
            tuned_fit, tuned_breakdown = self.fitness(tuned)
            improved = False
            if tuned_breakdown["hard_penalty_raw"] <= best_breakdown["hard_penalty_raw"]:
                if tuned_breakdown["soft_penalty_raw"] + 1e-9 < best_breakdown["soft_penalty_raw"]:
                    improved = True
                elif (
                    abs(tuned_breakdown["soft_penalty_raw"] - best_breakdown["soft_penalty_raw"]) <= 1e-9
                    and (
                        tuned_breakdown["weekend_off_lt4"] < best_breakdown["weekend_off_lt4"]
                        or tuned_breakdown["single_off_days"] < best_breakdown["single_off_days"]
                    )
                ):
                    improved = True
                elif tuned_fit < best_fit:
                    improved = True
            if improved:
                best, best_fit, best_breakdown = tuned, tuned_fit, tuned_breakdown
            else:
                break
            if best_breakdown["soft_penalty_raw"] <= 0.4:
                break
        return best
    # 一次完整 GA 搜尋
    def run_once(self, run_id: int = 1, callback=None) -> Tuple[np.ndarray, Dict[str, float], List[Dict[str, float]]]:
        self.reseed(self.config.seed + run_id * 1000)

        population = [self.generate_individual() for _ in range(self.config.population_size)]
        fitness_cache = self.evaluate_population(population)

        # 讓 run1 初始族群就先做一輪 micro tune，提升起跑點品質
        if self.config.use_cpsat_tune:
            order = np.argsort(fitness_cache)
            warmup_k = min(5, len(population))
            for idx in order[:warmup_k]:
                tuned = self.cpsat_micro_tune(population[idx].copy())
                tuned_fit, _ = self.fitness(tuned)
                if tuned_fit < fitness_cache[idx]:
                    population[idx] = tuned
                    fitness_cache[idx] = tuned_fit

        best_idx = int(np.argmin(fitness_cache))
        best = population[best_idx].copy()
        best_fit, best_breakdown = self.fitness(best)
        stagnation = 0
        history: List[Dict[str, float]] = []

        print(f"[Run {run_id}] Initial best fitness = {best_fit:.2f}")
        self.print_breakdown(run_id, 0, best_breakdown, self.current_mutation_rate(0, 0), stagnation)

        global_restart_used = 0

        for gen in range(1, self.config.generations + 1):
            order = np.argsort(fitness_cache)
            population = [population[idx] for idx in order]
            fitness_cache = [fitness_cache[idx] for idx in order]

            if gen % self.config.local_search_every == 0:
                top_k = min(self.config.local_search_top_k, len(population), self.config.elite_size)
                for elite_idx in range(top_k):
                    population[elite_idx] = self.local_search(population[elite_idx])
                fitness_cache = self.evaluate_population(population)
                order = np.argsort(fitness_cache)
                population = [population[idx] for idx in order]
                fitness_cache = [fitness_cache[idx] for idx in order]

            next_population = [population[i].copy() for i in range(self.config.elite_size)]

            while len(next_population) < self.config.population_size - self.config.random_immigrants:
                p1 = self.tournament_select(population, fitness_cache)
                p2 = self.tournament_select(population, fitness_cache)
                c1, c2 = self.crossover(p1, p2)
                c1 = self.mutate(c1, gen, stagnation)
                c2 = self.mutate(c2, gen, stagnation)
                next_population.extend([c1, c2])

            next_population = next_population[: self.config.population_size - self.config.random_immigrants]

            for _ in range(self.config.random_immigrants):
                next_population.append(self.generate_individual())

            population = next_population
            fitness_cache = self.evaluate_population(population)

            gen_best_idx = int(np.argmin(fitness_cache))
            gen_best = population[gen_best_idx]

            if self.config.use_cpsat_tune and (stagnation >= self.config.cpsat_tune_trigger or gen % self.config.cpsat_tune_every == 0):
                tuned = self.cpsat_micro_tune(gen_best.copy())
                tuned_fit, _ = self.fitness(tuned)
                raw_fit, _ = self.fitness(gen_best)
                if tuned_fit < raw_fit:
                    gen_best = tuned
                    population[gen_best_idx] = tuned
                    fitness_cache[gen_best_idx] = tuned_fit

            if self.config.use_global_cpsat_restart and stagnation >= self.config.global_cpsat_trigger:
                can_restart = global_restart_used < self.config.max_global_restarts_per_run
                if can_restart:
                    globally_tuned = self.cpsat_global_restart(gen_best.copy())
                    global_fit, _ = self.fitness(globally_tuned)
                    raw_fit, _ = self.fitness(gen_best)
                    if global_fit < raw_fit:
                        gen_best = globally_tuned
                        population[gen_best_idx] = globally_tuned
                        fitness_cache[gen_best_idx] = global_fit
                        stagnation = 0
                    global_restart_used += 1

            pre_polish_fit, pre_polish_breakdown = self.fitness(gen_best)
            if self.config.use_final_polishing:
                polish_ready = (
                    pre_polish_fit <= self.config.final_polish_threshold
                    or stagnation >= self.config.final_polish_trigger
                )
                if polish_ready and (gen % self.config.final_polish_every == 0) and pre_polish_breakdown["single_off_days"] > 0:
                    polished = self.cpsat_final_polish(gen_best.copy())
                    polished_fit, _ = self.fitness(polished)
                    if polished_fit < pre_polish_fit:
                        gen_best = polished
                        population[gen_best_idx] = polished
                        fitness_cache[gen_best_idx] = polished_fit

            # GA 後段才啟用 single-off repair，避免前期探索被提早修平
            if gen >= int(self.config.generations * 0.7):
                late_fixed = self.repair(gen_best.copy(), fix_single_off=True)
                late_fixed_fit, _ = self.fitness(late_fixed)
                raw_fit, _ = self.fitness(gen_best)
                if late_fixed_fit < raw_fit:
                    gen_best = late_fixed
                    population[gen_best_idx] = late_fixed
                    fitness_cache[gen_best_idx] = late_fixed_fit

            gen_best_fit, gen_breakdown = self.fitness(gen_best)

            if gen_best_fit < best_fit:
                best = gen_best.copy()
                best_fit = gen_best_fit
                best_breakdown = gen_breakdown
                stagnation = 0
            else:
                stagnation += 1

            current_mut = self.current_mutation_rate(gen, stagnation)

            history.append({
                "run_id": run_id,
                "generation": gen,
                "best_fitness": best_fit,
                "hard_penalty_raw": best_breakdown["hard_penalty_raw"],
                "soft_penalty_raw": best_breakdown["soft_penalty_raw"],
                "mutation_rate": current_mut,
                "stagnation": stagnation,
                "demand_mismatch": best_breakdown["demand_mismatch"],
                "preset_violation": best_breakdown["preset_violation"],
                "consecutive_6": best_breakdown["consecutive_6"],
                "transition_violation": best_breakdown["transition_violation"], # 合併後的鍵值
                "few_two_day_off_blocks": best_breakdown["few_two_day_off_blocks"],
                "monthly_off_lt9": best_breakdown["monthly_off_lt9"],
                "weekend_off_lt4": best_breakdown["weekend_off_lt4"],
                "single_off_days": best_breakdown["single_off_days"],
            })

            if gen == 1 or gen % self.config.print_every == 0:
                self.print_breakdown(run_id, gen, best_breakdown, current_mut, stagnation)

            if callback:
                callback(gen, self.config.generations, best_breakdown["total_fitness"])

            if best_breakdown["hard_penalty_raw"] <= 0.2 + 1e-9 and best_breakdown["soft_penalty_raw"] <= 0.4 + 1e-9:
                return best, best_breakdown, history

            if stagnation >= self.config.early_stop_rounds:
                print(f"[Run {run_id}] Early stop at generation {gen} (stagnation={stagnation})")
                break

        return best, best_breakdown, history
    # 從多次結果中挑最好的
    def run(self, callback=None) -> Tuple[np.ndarray, Dict[str, float], pd.DataFrame]:
        global_best = None
        global_breakdown = None
        global_best_fit = float("inf")
        all_history: List[Dict[str, float]] = []
        restart_summary: List[Dict[str, float]] = []

        total_gens = self.config.generations * self.config.restarts

        def _run_callback(gen, gen_total, fitness):
            if not callback:
                return
            overall_gen = (run_id - 1) * self.config.generations + gen
            callback(overall_gen, total_gens, fitness,
                     run_id=run_id, total_runs=self.config.restarts,
                     gen_in_run=gen, gens_per_run=self.config.generations)

        for run_id in range(1, self.config.restarts + 1):
            best, breakdown, history = self.run_once(run_id, callback=_run_callback)
            fit = breakdown["total_fitness"]
            all_history.extend(history)
            restart_summary.append({
                "run_id": run_id,
                "best_fitness": fit,
                "hard_penalty_raw": breakdown["hard_penalty_raw"],
                "soft_penalty_raw": breakdown["soft_penalty_raw"],
                "demand_mismatch": breakdown["demand_mismatch"],
                "preset_violation": breakdown["preset_violation"],
                "consecutive_6": breakdown["consecutive_6"],
                "transition_violation": breakdown["transition_violation"], # 修改此處
                "few_two_day_off_blocks": breakdown["few_two_day_off_blocks"],
                "monthly_off_lt9": breakdown["monthly_off_lt9"],
                "weekend_off_lt4": breakdown["weekend_off_lt4"],
                "single_off_days": breakdown["single_off_days"],
            })
            if fit < global_best_fit:
                global_best_fit = fit
                global_best = best.copy()
                global_breakdown = breakdown.copy()

            if global_best_fit <= self.config.target_fitness + 1e-9:
                print(f"[Overall] Target reached at run {run_id} (best={global_best_fit:.2f} <= {self.config.target_fitness:.2f})")
                break

        self.restart_df = pd.DataFrame(restart_summary)
        self.history_df = pd.DataFrame(all_history)
        return global_best, global_breakdown, self.history_df

    def to_output_df(self, schedule: np.ndarray) -> pd.DataFrame:
        out = self.engineer_df[[self.people_col, self.group_col]].copy()
        for d, col in enumerate(self.day_cols):
            out[col] = [INT_TO_SHIFT[int(x)] for x in schedule[:, d]]
        return out

    def save_output_csv(self, schedule: np.ndarray, out_csv: str = "Scheduling_Output.csv") -> None:
        self.to_output_df(schedule).to_csv(out_csv, index=False, encoding="utf-8-sig")

    def save_reports(self, best_breakdown: Dict[str, float], report_prefix: str = "Scheduling") -> None:
        penalty_df = pd.DataFrame([{"metric": k, "value": v} for k, v in best_breakdown.items()])
        penalty_df.to_csv(f"{report_prefix}_Penalty_Breakdown.csv", index=False, encoding="utf-8-sig")
        if hasattr(self, "restart_df"):
            self.restart_df.to_csv(f"{report_prefix}_Restart_Summary.csv", index=False, encoding="utf-8-sig")
        if hasattr(self, "history_df"):
            self.history_df.to_csv(f"{report_prefix}_Generation_History.csv", index=False, encoding="utf-8-sig")
        pd.DataFrame([asdict(self.config)]).to_csv(f"{report_prefix}_Config.csv", index=False, encoding="utf-8-sig")

    def save_output_xlsx(self, schedule: np.ndarray, best_breakdown: Dict[str, float], out_xlsx: str = "Scheduling_Output.xlsx") -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Schedule"

        fill_header = PatternFill("solid", fgColor="D9E2F3")
        fill_subheader = PatternFill("solid", fgColor="E2F0D9")
        fill_weekend = PatternFill("solid", fgColor="FFF2CC")
        fill_shift_o = PatternFill("solid", fgColor="FFFFFF")
        fill_shift_d = PatternFill("solid", fgColor="DDEBF7")
        fill_shift_e = PatternFill("solid", fgColor="FFF2CC")
        fill_shift_n = PatternFill("solid", fgColor="E4DFEC")
        fill_hard = PatternFill("solid", fgColor="FF6666")
        fill_preset = PatternFill("solid", fgColor="9E7CC1")
        fill_single = PatternFill("solid", fgColor="92D050")
        fill_emp_flag = PatternFill("solid", fgColor="FCE4D6")

        thin = Side(style="thin", color="999999")
        weekend_side = Side(style="thin", color="E3A500")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        white_font = Font(color="FFFFFF", bold=True)
        bold_font = Font(bold=True)

        ws["A1"] = "自動排班摘要"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = "總分"
        ws["B2"] = best_breakdown["total_fitness"]
        ws["A3"] = "硬懲罰"
        ws["B3"] = best_breakdown["hard_penalty_raw"]
        ws["A4"] = "軟懲罰"
        ws["B4"] = best_breakdown["soft_penalty_raw"]
        for cell in ["A2", "A3", "A4"]:
            ws[cell].font = bold_font
            ws[cell].fill = fill_header
        for cell in ["B2", "B3", "B4"]:
            ws[cell].fill = fill_subheader

        start_rule_row = 1
        rule_col = 5
        ws.cell(start_rule_row, rule_col).value = "題目規則懲罰"
        ws.cell(start_rule_row, rule_col).font = Font(bold=True, size=14)
        rule_items = [
            ("1.0 連續上班6天", best_breakdown["consecutive_6"]),
            ("1.0 班別轉換違規", best_breakdown["transition_violation"]), # 合併顯示
            ("0.2 違反預設班別", best_breakdown["preset_violation"]),
            ("0.1 連續休假次數<2", best_breakdown["few_two_day_off_blocks"]),
            ("0.1 月休<9", best_breakdown["monthly_off_lt9"]),
            ("0.1 周末休<4", best_breakdown["weekend_off_lt4"]),
            ("0.1 單休1日", best_breakdown["single_off_days"]),
        ]
        r = start_rule_row + 1
        for label, value in rule_items:
            ws.cell(r, rule_col).value = label
            ws.cell(r, rule_col + 1).value = value
            r += 1

        legend_col = 8
        ws.cell(1, legend_col).value = "圖例"
        ws.cell(1, legend_col).font = Font(bold=True, size=14)
        legends = [
            ("班別 O", fill_shift_o),
            ("班別 D", fill_shift_d),
            ("班別 E", fill_shift_e),
            ("班別 N", fill_shift_n),
            ("硬限制違規", fill_hard),
            ("違反預設班別", fill_preset),
            ("單休1日", fill_single),
            ("人員層級公平性提醒", fill_emp_flag),
        ]
        rr = 2
        for text, fill in legends:
            ws.cell(rr, legend_col).value = text
            ws.cell(rr, legend_col + 1).fill = fill
            ws.cell(rr, legend_col + 1).border = border
            rr += 1

        table_start_row = 12
        ws.cell(table_start_row, 1).value = "需求/人員"
        ws.cell(table_start_row + 1, 1).value = "D"
        ws.cell(table_start_row + 2, 1).value = "E"
        ws.cell(table_start_row + 3, 1).value = "N"

        for d, col in enumerate(self.day_cols):
            c = d + 3
            ws.cell(table_start_row, c).value = col
            ws.cell(table_start_row + 1, c).value = int(self.day_demand[d, 0])
            ws.cell(table_start_row + 2, c).value = int(self.day_demand[d, 1])
            ws.cell(table_start_row + 3, c).value = int(self.day_demand[d, 2])

            for rr in [table_start_row, table_start_row + 1, table_start_row + 2, table_start_row + 3]:
                ws.cell(rr, c).alignment = center

            ws.cell(table_start_row, c).fill = fill_weekend if self.weekend_flags[d] else fill_header

        for rr in range(table_start_row, table_start_row + 4):
            for cc in range(1, self.n_days + 3):
                ws.cell(rr, cc).border = border
                if rr > table_start_row and cc == 1:
                    ws.cell(rr, cc).fill = fill_subheader
                    ws.cell(rr, cc).font = bold_font

        data_start_row = table_start_row + 5
        ws.cell(data_start_row, 1).value = "人員"
        ws.cell(data_start_row, 2).value = "班別群組"
        ws.cell(data_start_row, 1).fill = fill_header
        ws.cell(data_start_row, 2).fill = fill_header
        ws.cell(data_start_row, 1).font = bold_font
        ws.cell(data_start_row, 2).font = bold_font
        ws.cell(data_start_row, 1).alignment = center
        ws.cell(data_start_row, 2).alignment = center

        for d, col in enumerate(self.day_cols):
            c = d + 3
            ws.cell(data_start_row, c).value = col
            ws.cell(data_start_row, c).alignment = center
            ws.cell(data_start_row, c).font = bold_font
            ws.cell(data_start_row, c).fill = fill_weekend if self.weekend_flags[d] else fill_header

        for i in range(self.n_people):
            rr = data_start_row + 1 + i
            ws.cell(rr, 1).value = self.people[i]
            ws.cell(rr, 2).value = INT_TO_SHIFT.get(self.groups[i], "")
            ws.cell(rr, 1).alignment = left_align
            ws.cell(rr, 2).alignment = center

            for d in range(self.n_days):
                c = d + 3
                shift_text = INT_TO_SHIFT[int(schedule[i, d])]
                cell = ws.cell(rr, c)
                cell.value = shift_text
                cell.alignment = center
                if shift_text == "O":
                    cell.fill = fill_shift_o
                elif shift_text == "D":
                    cell.fill = fill_shift_d
                elif shift_text == "E":
                    cell.fill = fill_shift_e
                else:
                    cell.fill = fill_shift_n
                cell.border = Border(left=thin, right=thin, top=weekend_side if self.weekend_flags[d] else thin, bottom=thin)

        violations = self.collect_violations(schedule)

        for i, d, reason in violations["preset_violation_cells"]:
            rr = data_start_row + 1 + i
            cc = 3 + d
            ws.cell(rr, cc).fill = fill_preset
            ws.cell(rr, cc).font = white_font
            ws.cell(rr, cc).comment = Comment(reason, "ChatGPT")

        for bucket in ["transition_cells", "consecutive_6_cells"]:
            for i, d, reason in violations[bucket]:
                rr = data_start_row + 1 + i
                cc = 3 + d
                ws.cell(rr, cc).fill = fill_hard
                ws.cell(rr, cc).font = white_font
                old_comment = ws.cell(rr, cc).comment.text + "\n" if ws.cell(rr, cc).comment else ""
                ws.cell(rr, cc).comment = Comment(old_comment + reason, "ChatGPT")

        for i, d, reason in violations["single_off_cells"]:
            rr = data_start_row + 1 + i
            cc = 3 + d
            current_fill = ws.cell(rr, cc).fill.fgColor.rgb
            if current_fill not in ["00FF6666", "FF6666", "009E7CC1", "9E7CC1"]:
                ws.cell(rr, cc).fill = fill_single
            old_comment = ws.cell(rr, cc).comment.text + "\n" if ws.cell(rr, cc).comment else ""
            ws.cell(rr, cc).comment = Comment(old_comment + reason, "ChatGPT")

        for i, reasons in violations["employee_flags"]:
            rr = data_start_row + 1 + i
            ws.cell(rr, 1).fill = fill_emp_flag
            ws.cell(rr, 1).comment = Comment("\n".join(reasons), "ChatGPT")

        ws.freeze_panes = "C18"
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 10
        for c in range(3, self.n_days + 3):
            letter = ws.cell(1, c).column_letter
            ws.column_dimensions[letter].width = 5

        for rr in range(1, data_start_row + self.n_people + 2):
            ws.row_dimensions[rr].height = 22

        ws2 = wb.create_sheet("Summary")
        ws2["A1"] = "Best penalty breakdown"
        ws2["A1"].font = Font(bold=True, size=14)
        r = 3
        for k, v in best_breakdown.items():
            ws2.cell(r, 1).value = k
            ws2.cell(r, 2).value = v
            r += 1
        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 16

        ws3 = wb.create_sheet("Config")
        ws3["A1"] = "GA Config"
        ws3["A1"].font = Font(bold=True, size=14)
        r = 3
        for k, v in asdict(self.config).items():
            ws3.cell(r, 1).value = k
            ws3.cell(r, 2).value = v
            r += 1
        ws3.column_dimensions["A"].width = 26
        ws3.column_dimensions["B"].width = 16

        wb.save(out_xlsx)


if __name__ == "__main__":
    # 這組參數是偏向「讓 run1 更有機會直接打到 0.6」的強化版
    config = GAConfig(
        target_fitness=0.6,
        population_size=100,
        generations=500,
        tournament_size=4,
        crossover_rate=0.90,
        mutation_rate=0.25,
        elite_size=4,
        random_immigrants=12,
        local_search_every=6,
        local_search_steps=12,
        stagnation_limit=20,
        restarts=1,
        seed=42,
        early_stop_rounds=120,
        local_search_top_k=3,
        print_every=20,
        use_cpsat_tune=True,
        cpsat_tune_every=6,
        cpsat_tune_trigger=4,
        cpsat_window_days=14,
        cpsat_max_employees=10,
        cpsat_time_limit=5.0,
        cpsat_num_workers=8,
        use_global_cpsat_restart=True,
        global_cpsat_trigger=8,
        global_cpsat_time_limit=12.0,
        global_cpsat_num_workers=8,
        global_cpsat_max_change_cells=80,
        global_cpsat_once_per_run=False,
        max_global_restarts_per_run=3,
        use_final_polishing=True,
        final_polish_threshold=1.2,
        final_polish_trigger=4,
        final_polish_every=2,
        final_polish_window_days=21,
        final_polish_max_employees=10,
        final_polish_time_limit=10.0,
        final_polish_num_workers=8,
        final_polish_rounds=4,
        final_polish_max_change_cells=100,
    )

    solver = SchedulerGA(
        engineer_csv="Engineer_List.csv",
        demand_csv="Shift_Demand.csv",
        config=config,
    )

    best_schedule, best_breakdown, history_df = solver.run()
    solver.save_output_csv(best_schedule, "Scheduling_Output.csv")
    solver.save_output_xlsx(best_schedule, best_breakdown, "Scheduling_Output.xlsx")
    solver.save_reports(best_breakdown, report_prefix="Scheduling")
    end = time.time()

    print(f"\nTotal runtime = {end-start:.2f} sec")
    print("\nBest penalty breakdown:")
    for k, v in best_breakdown.items():
        print(f"  {k}: {v}")

    print("\nSaved files:")
    print("  Scheduling_Output.csv")
    print("  Scheduling_Output.xlsx")
    print("  Scheduling_Penalty_Breakdown.csv")
    print("  Scheduling_Restart_Summary.csv")
    print("  Scheduling_Generation_History.csv")
    print("  Scheduling_Config.csv")

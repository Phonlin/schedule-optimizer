import io
import json
import os
import queue
import threading
import uuid

import pandas as pd
from flask import Flask, Response, redirect, render_template, request, url_for

from algorithms import REGISTRY

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "change-me-in-production")

DATE_COLS = [f"Date_{i}" for i in range(1, 31)]
_GROUP_COL_PATTERN = "班別群組"

# 儲存最近完成的排班結果（最多保留 20 筆）
_job_results: dict = {}
_MAX_JOBS = 20


# ──────────────────────────────────────────────
# CSV 解析
# ──────────────────────────────────────────────

def parse_engineer_list(file_storage) -> pd.DataFrame:
    try:
        df = pd.read_csv(file_storage)
    except Exception as exc:
        raise ValueError(f"無法解析 Engineer_List CSV：{exc}") from exc

    group_col = next(
        (c for c in df.columns if _GROUP_COL_PATTERN in str(c)), None
    )
    if group_col is None:
        raise ValueError(
            "Engineer_List CSV 缺少「班別群組」欄位。"
            f"偵測到的欄位：{df.columns.tolist()}"
        )
    if "人員" not in df.columns:
        raise ValueError(
            "Engineer_List CSV 缺少「人員」欄位。"
            f"偵測到的欄位：{df.columns.tolist()}"
        )

    present_dates = [c for c in DATE_COLS if c in df.columns]
    if not present_dates:
        raise ValueError("Engineer_List CSV 未找到 Date_1~Date_30 任何日期欄位。")

    for col in present_dates:
        df[col] = df[col].fillna("").astype(str).str.strip().str.upper()

    df["班別群組"] = df[group_col].fillna("").astype(str).str.strip().str.upper()
    df["primary_group"] = df["班別群組"].str[:1]
    df["backup_groups"] = df["班別群組"].str[1:].apply(lambda s: list(s) if s else [])

    for col in DATE_COLS:
        if col not in df.columns:
            df[col] = ""

    return df


def parse_shift_demand(file_storage) -> pd.DataFrame:
    try:
        df = pd.read_csv(file_storage)
    except Exception as exc:
        raise ValueError(f"無法解析 Shift_Demand CSV：{exc}") from exc

    df.columns = df.columns.str.strip()
    required = ["Date", "Day", "Afternoon", "Night"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(
            f"Shift_Demand CSV 缺少欄位：{', '.join(missing)}。"
            f"偵測到的欄位：{df.columns.tolist()}"
        )

    if "IfWeekend" not in df.columns:
        df["IfWeekend"] = ""

    for col in ["Day", "Afternoon", "Night"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)

    df["IfWeekend"] = (
        df["IfWeekend"].fillna("").astype(str).str.strip().str.upper() == "Y"
    )
    return df[["Date", "IfWeekend", "Day", "Afternoon", "Night"]]


# ──────────────────────────────────────────────
# 序列化工具
# ──────────────────────────────────────────────

def _grid_to_json(grid_df: pd.DataFrame) -> str:
    return grid_df.to_json(orient="split", force_ascii=False)


def _json_to_grid(json_str: str) -> pd.DataFrame:
    data = json.loads(json_str)
    df = pd.DataFrame(data["data"], index=data["index"], columns=data["columns"])
    df.index.name = "人員"
    return df


# ──────────────────────────────────────────────
# 違規分析
# ──────────────────────────────────────────────

_WORK_SHIFTS = {"D", "E", "N"}
_INVALID_TRANS = {("N", "D"), ("N", "E"), ("E", "D")}
_GROUP_TO_PS = {"D": "D", "E": "E", "N": "N"}

# 違規優先順序（越前面越嚴重，多重違規顯示最高優先）
_VIOL_PRIORITY = ["transition", "consec", "default_shift", "single_off"]

# 違規 → CSS class
_VIOL_CSS = {
    "transition":    "viol-transition",
    "consec":        "viol-consec",
    "default_shift": "viol-default-shift",
    "single_off":    "viol-single-off",
}

# 違規 → 中文說明
_VIOL_LABEL = {
    "transition":    "班別轉換違規 (N→D / N→E / E→D)",
    "consec":        "連續上班 ≥6 天",
    "default_shift": "違反預設班別",
    "single_off":    "單休1日",
    "monthly_off":   "月休 <9 天",
    "weekend_off":   "周末休 <4 天",
    "rest_blocks":   "連續休假段數 <2",
}

# 無違規時的班別 CSS class
_SHIFT_CSS = {
    "D": "shift-D", "E": "shift-E", "N": "shift-N",
    "O": "shift-O", "": "shift-empty",
}

# CSS class → (背景 hex, 字體 hex)  供 XLSX 使用
_XLSX_BG_FG: dict[str, tuple[str, str]] = {
    "shift-D":            ("CFE2FF", "084298"),
    "shift-E":            ("D1E7DD", "0A3622"),
    "shift-N":            ("E2D9F3", "432874"),
    "shift-O":            ("F8D7DA", "842029"),
    "shift-empty":        ("FFFFFF", "ADB5BD"),
    "viol-transition":    ("FF5252", "FFFFFF"),
    "viol-consec":        ("FF9800", "FFFFFF"),
    "viol-default-shift": ("FFD600", "333333"),
    "viol-single-off":    ("40C4FF", "01579B"),
}


def _rl(indices: list[int]) -> list[int]:
    """計算連續索引段的各段長度。"""
    if not indices:
        return []
    runs, length = [], 1
    for i in range(1, len(indices)):
        if indices[i] == indices[i - 1] + 1:
            length += 1
        else:
            runs.append(length)
            length = 1
    runs.append(length)
    return runs


def analyze_violations(grid_df: pd.DataFrame, demand_df: pd.DataFrame) -> tuple:
    """
    分析排班結果中的每格與每列違規情況。

    Returns
    -------
    cell_classes : {eng: {date: css_class}}   每格最終應套用的 CSS class
    cell_titles  : {eng: {date: tooltip_str}} 違規說明（hover tooltip）
    row_flags    : {eng: [violation_codes]}   列層級違規代碼清單
    """
    dates = [c for c in grid_df.columns if c != "班別群組"]
    engineers = grid_df.index.tolist()

    is_weekend = {row["Date"]: bool(row["IfWeekend"]) for _, row in demand_df.iterrows()}

    cell_viols:  dict[str, dict[str, list[str]]] = {
        eng: {d: [] for d in dates} for eng in engineers
    }
    row_flags: dict[str, list[str]] = {eng: [] for eng in engineers}

    for eng in engineers:
        row = grid_df.loc[eng]
        group_str = str(row.get("班別群組", "D")).strip().upper()
        primary = _GROUP_TO_PS.get(group_str[:1], "D")
        schedule = {d: str(row.get(d, "")).strip() for d in dates}

        work_seq = [
            (i, d, schedule[d])
            for i, d in enumerate(dates)
            if schedule[d] in _WORK_SHIFTS
        ]

        # ── 連續上班 ≥6 天 ──────────────────────────────────────────
        if work_seq:
            consec = 1
            for k in range(1, len(work_seq)):
                if work_seq[k][0] == work_seq[k - 1][0] + 1:
                    consec += 1
                    if consec >= 6:
                        cell_viols[eng][work_seq[k][1]].append("consec")
                else:
                    consec = 1

        # ── 班別轉換違規 ─────────────────────────────────────────────
        for k in range(1, len(work_seq)):
            pi, _, ps = work_seq[k - 1]
            ci, cd, cs = work_seq[k]
            if ci == pi + 1 and (ps, cs) in _INVALID_TRANS:
                cell_viols[eng][cd].append("transition")

        # ── 違反預設班別 ─────────────────────────────────────────────
        for _, d, s in work_seq:
            if s != primary:
                cell_viols[eng][d].append("default_shift")

        # ── 單休1日 ──────────────────────────────────────────────────
        for i, d in enumerate(dates):
            if schedule[d] not in _WORK_SHIFTS:
                prev_work = i > 0 and schedule[dates[i - 1]] in _WORK_SHIFTS
                next_work = i < len(dates) - 1 and schedule[dates[i + 1]] in _WORK_SHIFTS
                if prev_work and next_work:
                    cell_viols[eng][d].append("single_off")

        # ── 月休 <9（列層級）───────────────────────────────────────
        off_cnt = sum(1 for d in dates if schedule[d] not in _WORK_SHIFTS)
        if off_cnt < 9:
            row_flags[eng].append("monthly_off")

        # ── 周末休 <4（列層級）─────────────────────────────────────
        wkend_off = sum(
            1 for d in dates
            if schedule[d] not in _WORK_SHIFTS and is_weekend.get(d, False)
        )
        if wkend_off < 4:
            row_flags[eng].append("weekend_off")

        # ── 連續休假段數 <2（列層級）────────────────────────────────
        off_idx = [i for i, d in enumerate(dates) if schedule[d] not in _WORK_SHIFTS]
        if len(_rl(off_idx)) < 2:
            row_flags[eng].append("rest_blocks")

    # ── 將違規代碼轉換為 CSS class 與 tooltip ──────────────────────
    cell_classes: dict[str, dict[str, str]] = {}
    cell_titles:  dict[str, dict[str, str]] = {}

    for eng in engineers:
        cell_classes[eng] = {}
        cell_titles[eng]  = {}
        schedule = {d: str(grid_df.loc[eng].get(d, "")).strip() for d in dates}

        for d in dates:
            viols = cell_viols[eng][d]
            # 選出最高優先違規的 CSS class
            css = next(
                (_VIOL_CSS[v] for v in _VIOL_PRIORITY if v in viols),
                _SHIFT_CSS.get(schedule[d], "shift-empty"),   # 無違規 → 班別色
            )
            cell_classes[eng][d] = css
            cell_titles[eng][d]  = " ｜ ".join(_VIOL_LABEL[v] for v in viols)

    return cell_classes, cell_titles, row_flags, cell_viols


def _algorithms_list():
    return [{"key": k, "name": v.name} for k, v in REGISTRY.items()]


# ──────────────────────────────────────────────
# Flask 路由
# ──────────────────────────────────────────────

@app.route("/")
def index():
    return render_template(
        "index.html",
        error=None,
        grid=None,
        dates=None,
        penalty=None,
        job_id=None,
        algorithms=_algorithms_list(),
        date_cols=DATE_COLS,
    )


@app.route("/run", methods=["POST"])
def run_job():
    """接受 CSV 上傳，以 SSE 串流回傳 GA 演化進度，完成後通知 job_id。"""
    # 先讀取檔案（請求結束後不能再讀）
    shift_file = request.files.get("shift_csv")
    staff_file = request.files.get("staff_csv")
    algo_key = request.form.get("algorithm", "genetic")

    def _sse_error(msg: str):
        def gen():
            yield f'data: {json.dumps({"type": "error", "message": msg}, ensure_ascii=False)}\n\n'
        return Response(gen(), mimetype="text/event-stream",
                        headers={"Cache-Control": "no-cache"})

    if not shift_file or shift_file.filename == "":
        return _sse_error("請上傳班別需求 CSV（Shift_Demand）。")
    if not staff_file or staff_file.filename == "":
        return _sse_error("請上傳工程師清單 CSV（Engineer_List）。")
    if algo_key not in REGISTRY:
        return _sse_error(f"不支援的算法：{algo_key}")

    # 讀取到 BytesIO（執行緒安全）
    shift_bytes = io.BytesIO(shift_file.read())
    staff_bytes = io.BytesIO(staff_file.read())

    try:
        demand_df = parse_shift_demand(shift_bytes)
        staff_df = parse_engineer_list(staff_bytes)
    except ValueError as exc:
        return _sse_error(str(exc))

    job_id = uuid.uuid4().hex[:8]
    progress_q: queue.Queue = queue.Queue()

    def ga_worker():
        def on_progress(gen, total, fitness, **kwargs):
            msg = {
                "type": "progress",
                "gen": gen,
                "total": total,
                "fitness": fitness,
            }
            msg.update(kwargs)
            progress_q.put(msg)

        try:
            scheduler = REGISTRY[algo_key]
            grid_df, penalty = scheduler.run(staff_df, demand_df, callback=on_progress)

            # 儲存結果，超過上限時移除最舊的
            _job_results[job_id] = {
                "grid_df": grid_df,
                "penalty": penalty,
                "demand_df": demand_df,   # 供違規分析使用
                "staff_df": staff_df.copy(),  # Gurobi 參考求解需預排與群組
            }
            while len(_job_results) > _MAX_JOBS:
                del _job_results[next(iter(_job_results))]

            progress_q.put({"type": "done", "job_id": job_id, "penalty": penalty})
        except Exception as exc:  # noqa: BLE001
            progress_q.put({"type": "error", "message": str(exc)})

    threading.Thread(target=ga_worker, daemon=True).start()

    def event_stream():
        while True:
            try:
                msg = progress_q.get(timeout=600)   # 10 分鐘逾時
            except queue.Empty:
                yield 'data: {"type":"error","message":"計算逾時，請重試。"}\n\n'
                break
            yield f"data: {json.dumps(msg, ensure_ascii=False)}\n\n"
            if msg["type"] in ("done", "error"):
                break

    return Response(
        event_stream(),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


def _sse_verify_error(msg: str):
    def gen():
        yield f'data: {json.dumps({"type": "error", "message": msg}, ensure_ascii=False)}\n\n'

    return Response(
        gen(),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.route("/verify_gurobi/<job_id>", methods=["POST"])
def verify_gurobi(job_id):
    """以 Gurobi MIP 求參考最佳，與當次演算法懲罰對照（SSE）。"""
    result = _job_results.get(job_id)
    if not result:
        return _sse_verify_error("結果已過期或不存在，請重新排班。")
    if "staff_df" not in result:
        return _sse_verify_error("此結果無法驗證（缺少原始工程師資料），請重新排班。")

    progress_q: queue.Queue = queue.Queue()

    def verify_worker():
        try:
            from gurobi_reference import (
                GurobiReferenceError,
                breakdown_for_json,
                diff_grid_vs_schedule,
                solve_reference_mip,
            )
        except ImportError as exc:
            progress_q.put(
                {
                    "type": "error",
                    "message": f"無法載入 Gurobi 模組：{exc}",
                }
            )
            return

        try:
            staff_df = result["staff_df"]
            demand_df = result["demand_df"]
            grid_df = result["grid_df"]
            penalty = result["penalty"]

            def on_log(text: str) -> None:
                progress_q.put({"type": "progress", "message": text, "phase": "gurobi"})

            on_log("啟動 Gurobi 參考求解…")
            sol = solve_reference_mip(staff_df, demand_df, log=on_log)

            diffs = diff_grid_vs_schedule(
                grid_df,
                sol["schedule"],
                sol["engineers"],
                sol["date_cols"],
            )

            algo_total = float(penalty.get("total", 0))
            gobj = float(sol["gurobi_obj"])
            delta = round(algo_total - gobj, 4)
            if gobj > 1e-9:
                delta_pct = round((algo_total - gobj) / gobj * 100, 2)
            else:
                delta_pct = None

            progress_q.put(
                {
                    "type": "done",
                    "phase": "gurobi",
                    "gurobi_obj": gobj,
                    "gurobi_status": sol["gurobi_status"],
                    "mip_gap": sol["mip_gap"],
                    "algo_total": algo_total,
                    "delta": delta,
                    "delta_pct": delta_pct,
                    "breakdown_gurobi": breakdown_for_json(sol["breakdown"]),
                    "algo_penalty": {
                        "total": penalty.get("total", 0),
                        "consecutive_6": penalty.get("consecutive_6", 0),
                        "transition": penalty.get("transition", 0),
                        "default_shift": penalty.get("default_shift", 0),
                        "rest_blocks": penalty.get("rest_blocks", 0),
                        "monthly_off": penalty.get("monthly_off", 0),
                        "weekend_off": penalty.get("weekend_off", 0),
                        "single_off": penalty.get("single_off", 0),
                        "demand": penalty.get("demand", 0),
                    },
                    "diff_total": len(diffs),
                    "diff_cells": diffs[:200],
                    "diff_truncated": len(diffs) > 200,
                }
            )
        except GurobiReferenceError as exc:
            progress_q.put({"type": "error", "message": str(exc)})
        except Exception as exc:  # noqa: BLE001
            progress_q.put({"type": "error", "message": str(exc)})

    threading.Thread(target=verify_worker, daemon=True).start()

    def event_stream():
        while True:
            try:
                msg = progress_q.get(timeout=600)
            except queue.Empty:
                yield (
                    'data: {"type":"error","message":"Gurobi 驗證逾時，請重試。"}\n\n'
                )
                break
            yield f"data: {json.dumps(msg, ensure_ascii=False)}\n\n"
            if msg["type"] in ("done", "error"):
                break

    return Response(
        event_stream(),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.route("/result/<job_id>")
def show_result(job_id):
    result = _job_results.get(job_id)
    if not result:
        return redirect(url_for("index"))

    grid_df   = result["grid_df"]
    penalty   = result["penalty"]
    demand_df = result["demand_df"]
    dates     = [c for c in grid_df.columns if c != "班別群組"]

    cell_classes, cell_titles, row_flags, _ = analyze_violations(grid_df, demand_df)

    is_weekend = {
        row["Date"]: bool(row["IfWeekend"])
        for _, row in demand_df.iterrows()
    }

    return render_template(
        "index.html",
        error=None,
        grid=grid_df,
        dates=dates,
        penalty=penalty,
        job_id=job_id,
        algorithms=_algorithms_list(),
        date_cols=DATE_COLS,
        cell_classes=cell_classes,
        cell_titles=cell_titles,
        row_flags=row_flags,
        is_weekend=is_weekend,
    )


@app.route("/download/csv/<job_id>")
def download_csv(job_id):
    result = _job_results.get(job_id)
    if not result:
        return "結果已過期，請重新排班。", 404

    grid_df   = result["grid_df"]
    demand_df = result["demand_df"]
    dates = [c for c in grid_df.columns if c != "班別群組"]

    out_df = grid_df.copy().reset_index()   # 讓「人員」成為第一欄欄頭
    out_df.rename(columns={"index": "人員"}, inplace=True)
    for col in dates:
        out_df[col] = out_df[col].replace("", "O").fillna("O")

    buf = io.StringIO()
    out_df.to_csv(buf, index=False)         # index=False 避免多餘的索引欄

    # ── 懲罰分數 ──────────────────────────────────────────────────
    penalty = result["penalty"]
    _PENALTY_LABELS = [
        ("total",         "總懲罰分數"),
        ("demand",        "需求未達標"),
        ("consecutive_6", "連續上班 ≥6 天"),
        ("transition",    "班別轉換違規"),
        ("default_shift", "違反預設班別"),
        ("rest_blocks",   "連續休假段 <2"),
        ("monthly_off",   "月休 <9 天"),
        ("weekend_off",   "周末休 <4 天"),
        ("single_off",    "單休1日"),
    ]
    buf.write("\n懲罰分數,數值\n")
    for key, label in _PENALTY_LABELS:
        buf.write(f"{label},{penalty.get(key, 0)}\n")

    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": "attachment; filename=schedule.csv"},
    )


@app.route("/download/xlsx/<job_id>")
def download_xlsx(job_id):
    result = _job_results.get(job_id)
    if not result:
        return "結果已過期，請重新排班。", 404

    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment

    grid_df   = result["grid_df"]
    demand_df = result["demand_df"]
    penalty   = result["penalty"]
    dates = [c for c in grid_df.columns if c != "班別群組"]
    engineers = grid_df.index.tolist()

    cell_classes, cell_titles, row_flags, cell_viols = analyze_violations(
        grid_df, demand_df
    )

    wb = Workbook()

    # ── Sheet 1：排班表（彩色）────────────────────────────────────
    ws = wb.active
    ws.title = "排班表"

    # 標頭
    header = ["人員", "班別群組"] + dates + ["月", "末", "段"]

    # 判斷周末來設定休假色
    is_weekend = {
        row["Date"]: bool(row["IfWeekend"]) 
        for _, row in demand_df.iterrows()
    }
    
    # 標頭顏色
    for ci, h in enumerate(header, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        
        # 放假見紅
        if is_weekend.get(h, False):
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        else:
            cell.fill = PatternFill(start_color="343A40", end_color="343A40", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    _ROW_VIOL_COLS = {
        "monthly_off": (len(dates) + 3, "F48FB1", "880E4F"),
        "weekend_off": (len(dates) + 4, "CE93D8", "4A148C"),
        "rest_blocks": (len(dates) + 5, "FFAB91", "BF360C"),
    }

    # 資料列
    for ri, eng in enumerate(engineers, 2):
        # 人員
        ws.cell(row=ri, column=1, value=eng).font = Font(bold=True)
        # 班別群組
        ws.cell(row=ri, column=2, value=str(grid_df.loc[eng].get("班別群組", "")))

        # 日期格
        for ci, d in enumerate(dates, 3):
            val = str(grid_df.loc[eng].get(d, "")) or "O"
            css = cell_classes.get(eng, {}).get(d, "shift-empty")
            bg, fg = _XLSX_BG_FG.get(css, ("FFFFFF", "000000"))

            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            cell.font = Font(color=fg)
            cell.alignment = Alignment(horizontal="center")

        # 列違規指示格（月/末/段，有違規填色，無違規留白）
        rf = row_flags.get(eng, [])
        for vk, (col_idx, vbg, vfg) in _ROW_VIOL_COLS.items():
            ic = ws.cell(row=ri, column=col_idx, value="")
            if vk in rf:
                ic.fill = PatternFill(start_color=vbg, end_color=vbg, fill_type="solid")

    # 欄寬
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    for ci in range(3, len(dates) + 3):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = 8
    for ci in range(len(dates) + 3, len(dates) + 6):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = 4

    # ── Sheet 1：懲罰分數 ─────────────────────────────────────────
    penalty_start = len(engineers) + 3
    ws.cell(row=penalty_start, column=1, value="懲罰分數").font = Font(bold=True, size=11)
    ws.cell(row=penalty_start, column=1).fill = PatternFill(
        start_color="343A40", end_color="343A40", fill_type="solid"
    )
    ws.cell(row=penalty_start, column=1).font = Font(color="FFFFFF", bold=True, size=11)
    ws.cell(row=penalty_start, column=2, value="數值").fill = PatternFill(
        start_color="343A40", end_color="343A40", fill_type="solid"
    )
    ws.cell(row=penalty_start, column=2).font = Font(color="FFFFFF", bold=True)

    _PENALTY_ROWS = [
        ("total",         "總懲罰分數",      "FF5252", "FFFFFF"),
        ("demand",        "需求未達標",      "FF9800", "FFFFFF"),
        ("consecutive_6", "連續上班 ≥6 天",  "FFD600", "333333"),
        ("transition",    "班別轉換違規",    "FF5252", "FFFFFF"),
        ("default_shift", "違反預設班別",    "FFD600", "333333"),
        ("rest_blocks",   "連續休假段 <2",   "FFAB91", "BF360C"),
        ("monthly_off",   "月休 <9 天",      "F48FB1", "880E4F"),
        ("weekend_off",   "周末休 <4 天",    "CE93D8", "4A148C"),
        ("single_off",    "單休1日",         "40C4FF", "01579B"),
    ]
    for i, (key, label, pbg, pfg) in enumerate(_PENALTY_ROWS):
        r = penalty_start + 1 + i
        lc = ws.cell(row=r, column=1, value=label)
        lc.fill = PatternFill(start_color=pbg, end_color=pbg, fill_type="solid")
        lc.font = Font(color=pfg)
        val = penalty.get(key, 0)
        vc = ws.cell(row=r, column=2, value=val)
        vc.alignment = Alignment(horizontal="right")
        if val and float(val) > 0:
            vc.font = Font(color="DC3545", bold=True)

    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width, 10)

    # ── Sheet 1 底部圖例 ──────────────────────────────────────────
    legend_start = penalty_start + len(_PENALTY_ROWS) + 2
    ws.cell(row=legend_start, column=1, value="圖例").font = Font(bold=True, size=11)
    _legend_items = [
        ("D", "D 日班",              "CFE2FF", "084298"),
        ("E", "E 午班",              "D1E7DD", "0A3622"),
        ("N", "N 晚班",              "E2D9F3", "432874"),
        ("O", "O 休假",              "F8D7DA", "842029"),
        ("",  "",                    "FFFFFF", "000000"),
        ("轉", "班別轉換違規 (N→D/N→E/E→D)", "FF5252", "FFFFFF"),
        ("連", "連續上班 ≥6 天",      "FF9800", "FFFFFF"),
        ("預", "違反預設班別",        "FFD600", "333333"),
        ("單", "單休1日",             "40C4FF", "01579B"),
        ("",  "",                    "FFFFFF", "000000"),
        ("月", "月欄 ■ 月休 <9 天",   "F48FB1", "880E4F"),
        ("末", "末欄 ■ 周末休 <4 天", "CE93D8", "4A148C"),
        ("段", "段欄 ■ 連續休假段 <2","FFAB91", "BF360C"),
    ]
    for i, (short, desc, bg, fg) in enumerate(_legend_items):
        r = legend_start + 1 + i
        swatch = ws.cell(row=r, column=1, value=short)
        swatch.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        swatch.font = Font(color=fg, bold=True)
        swatch.alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=desc).font = Font(size=10)

    # ── Sheet 2：違規清單 ─────────────────────────────────────────
    ws2 = wb.create_sheet("違規清單")
    ws2_header = ["人員", "日期", "違規類型"]
    for ci, h in enumerate(ws2_header, 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.fill = PatternFill(start_color="343A40", end_color="343A40", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)

    # 違規顏色
    _VIOL_BG = {
        "transition":    "FF5252", "consec":        "FF9800",
        "default_shift": "FFD600", "single_off":    "40C4FF",
        "monthly_off":   "F48FB1", "weekend_off":   "CE93D8",
        "rest_blocks":   "FFAB91",
    }
    _VIOL_FG = {
        "transition": "FFFFFF", "consec": "FFFFFF", "default_shift": "333333",
        "single_off": "01579B", "monthly_off": "880E4F",
        "weekend_off": "4A148C", "rest_blocks": "BF360C",
    }

    ws2_row = 2
    for eng in engineers:
        for d in dates:
            for v in cell_viols.get(eng, {}).get(d, []):
                r = ws2_row
                ws2.cell(row=r, column=1, value=eng)
                ws2.cell(row=r, column=2, value=d)
                vc = ws2.cell(row=r, column=3, value=_VIOL_LABEL[v])
                bg = _VIOL_BG.get(v, "FFFFFF")
                fg = _VIOL_FG.get(v, "000000")
                vc.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                vc.font = Font(color=fg, bold=True)
                ws2_row += 1

        for v in row_flags.get(eng, []):
            r = ws2_row
            ws2.cell(row=r, column=1, value=eng)
            ws2.cell(row=r, column=2, value="（整月）")
            vc = ws2.cell(row=r, column=3, value=_VIOL_LABEL[v])
            bg = _VIOL_BG.get(v, "FFFFFF")
            fg = _VIOL_FG.get(v, "000000")
            vc.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            vc.font = Font(color=fg, bold=True)
            ws2_row += 1

    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 34

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=schedule.xlsx"},
    )


if __name__ == "__main__":
    app.run(
        debug=True,
        host="0.0.0.0",
        port=int(os.environ.get("PORT", "5000")),
        threaded=True,
    )
